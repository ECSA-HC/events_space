import io
import math
from typing import Annotated
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.database import get_db
from dependencies.auth_dependency import Auth, get_current_user
from models.models import Registration, User, Event

router = APIRouter()

user_dependency = Annotated[dict, Depends(get_current_user)]


def get_auth_dep(db: Session = Depends(get_db)) -> Auth:
    return Auth(db)


def _serialize_reg(r: Registration) -> dict:
    user = r.user
    profile = user.user_profile[0] if user.user_profile else None
    return {
        "id": r.id,
        "event_id": r.event_id,
        "event": r.events.event if r.events else None,
        "user_id": r.user_id,
        "firstname": user.firstname,
        "lastname": user.lastname,
        "email": user.email,
        "phone": user.phone,
        "title": profile.title if profile else "",
        "organisation": profile.organisation if profile else "",
        "country": profile.country.country if profile and profile.country else "",
        "participation_role": r.participation_role.name if r.participation_role else "",
        "paid": r.paid,
        "registered_at": r.registered_at,
    }


@router.get("/participant_status/{registration_id}")
def participant_status(registration_id: int, db: Session = Depends(get_db)):
    from models.models import EventAttendance, Payment
    registration = (
        db.query(Registration)
        .options(
            joinedload(Registration.user),
            joinedload(Registration.payment),
            joinedload(Registration.event_attendance),
        )
        .filter(Registration.id == registration_id)
        .first()
    )
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    user = registration.user
    payment = registration.payment
    attendance_records = registration.event_attendance
    attendance_dates = [att.attendance_date.date() for att in attendance_records]

    return {
        "registration_id": registration.id,
        "participation_role": (
            registration.participation_role.name
            if registration.participation_role
            else None
        ),
        "paid": registration.paid,
        "user": {
            "id": user.id,
            "firstname": user.firstname,
            "lastname": user.lastname,
            "phone": user.phone,
            "email": user.email,
        },
        "payment": (
            {
                "id": payment.id if payment else None,
                "amount": getattr(payment, "amount", None),
                "status": getattr(payment, "status", None),
                "paid_at": getattr(payment, "paid_at", None),
            }
            if payment
            else None
        ),
        "attendance": {
            "count": len(attendance_records),
            "dates": attendance_dates,
        },
    }


@router.get("/")
async def list_registrations(
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
    event_id: int = Query(default=None),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=10),
    search: str = Query(default=""),
    paid: str = Query(default="all"),
):
    auth_dependency.secure_access("VIEW_REGISTRATIONS", current_user["user_id"])

    q = (
        db.query(Registration)
        .join(Registration.user)
        .options(
            joinedload(Registration.user).joinedload(User.user_profile),
            joinedload(Registration.events),
        )
        .filter(Registration.deleted_at == None)
    )

    if event_id:
        q = q.filter(Registration.event_id == event_id)

    if paid != "all":
        q = q.filter(Registration.paid == (paid == "true"))

    if search:
        term = f"%{search}%"
        q = q.filter(
            or_(
                User.firstname.ilike(term),
                User.lastname.ilike(term),
                User.email.ilike(term),
                User.phone.ilike(term),
            )
        )

    total = q.count()
    registrations = q.order_by(Registration.registered_at.desc()).offset(skip).limit(limit).all()
    pages = math.ceil(total / limit) if limit else 1

    return {
        "pages": pages,
        "total": total,
        "data": [_serialize_reg(r) for r in registrations],
    }


@router.get("/export")
async def export_registrations(
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
    event_id: int = Query(default=None),
    paid: str = Query(default="all"),
    search: str = Query(default=""),
):
    auth_dependency.secure_access("EXPORT_REGISTRATIONS", current_user["user_id"])

    q = (
        db.query(Registration)
        .join(Registration.user)
        .options(
            joinedload(Registration.user).joinedload(User.user_profile),
            joinedload(Registration.events),
            joinedload(Registration.payment),
        )
        .filter(Registration.deleted_at == None)
    )

    if event_id:
        q = q.filter(Registration.event_id == event_id)

    if paid != "all":
        q = q.filter(Registration.paid == (paid == "true"))

    if search:
        term = f"%{search}%"
        q = q.filter(
            or_(
                User.firstname.ilike(term),
                User.lastname.ilike(term),
                User.email.ilike(term),
            )
        )

    registrations = q.order_by(Registration.registered_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations"

    header_fill = PatternFill("solid", start_color="0095B6")
    alt_fill = PatternFill("solid", start_color="E8F4F8")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = [
        "#", "Title", "First Name", "Last Name", "Email", "Phone",
        "Organisation", "Country", "Event", "Participation Role",
        "Paid", "Registered At",
    ]
    ws.row_dimensions[1].height = 22
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws.freeze_panes = "A2"

    for ri, r in enumerate(registrations, 2):
        use_fill = alt_fill if ri % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        user = r.user
        profile = user.user_profile[0] if user.user_profile else None
        row = [
            r.id,
            profile.title if profile else "",
            user.firstname,
            user.lastname,
            user.email,
            user.phone,
            profile.organisation if profile else "",
            profile.country.country if profile and profile.country else "",
            r.events.event if r.events else "",
            r.participation_role.name if r.participation_role else "",
            "Yes" if r.paid else "No",
            r.registered_at.strftime("%d %b %Y %H:%M") if r.registered_at else "",
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            cell.font = body_font
            cell.fill = use_fill
            cell.alignment = left

    col_widths = [6, 8, 18, 18, 30, 16, 28, 20, 36, 18, 8, 18]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "registrations_export.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
