import io
from typing import Annotated, Optional
from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Request, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.database import get_db
from dependencies.auth_dependency import Auth, get_current_user, get_optional_current_user
from dependencies.dependency import Dependency
import utils.mailer_util as mailer_util
from models.models import Abstract, AbstractAuthor, AbstractReviewer, AbstractReview, AbstractStatus, User, UserRole, Role, EventTrack
from datetime import datetime
from schemas.events_space import (
    AbstractSubmitSchema, AbstractUpdateSchema,
    AssignReviewerSchema, BulkAssignReviewerSchema, AbstractReviewSchema,
    CreateReviewerSchema,
)

router = APIRouter()

user_dependency = Annotated[dict, Depends(get_current_user)]


def get_auth_dep(db: Session = Depends(get_db)) -> Auth:
    return Auth(db)


def get_dependency(db: Session = Depends(get_db)) -> Dependency:
    return Dependency(db)


def _serialize_abstract(a: Abstract):
    return {
        "id": a.id,
        "event_id": a.event_id,
        "event": a.event.event if a.event else None,
        "title": a.title,
        "abstract_text": a.abstract_text,
        "keywords": a.keywords,
        "track": a.track,
        "track_id": a.track_id,
        "track_code": a.event_track.code if a.event_track else None,
        "track_title": a.event_track.title if a.event_track else None,
        "track_theme": a.event_track.theme if a.event_track else None,
        "presentation_type": a.presentation_type.value if a.presentation_type else None,
        "status": a.status.value if a.status else None,
        "word_count": a.word_count,
        "submitted_by": a.submitted_by,
        "submitter_name": f"{a.submitter.firstname} {a.submitter.lastname}" if a.submitter else (f"{a.authors[0].firstname} {a.authors[0].lastname}" if a.authors else None),
        "submitter_email": a.submitter.email if a.submitter else (a.authors[0].email if a.authors else None),
        "acceptance_notified_at": a.acceptance_notified_at,
        "created_at": a.created_at,
        "updated_at": a.updated_at,
        "authors": [
            {
                "id": au.id,
                "firstname": au.firstname,
                "lastname": au.lastname,
                "email": au.email,
                "affiliation": au.affiliation,
                "country": au.country,
                "is_presenting": au.is_presenting,
                "author_order": au.author_order,
            }
            for au in a.authors
        ],
        "reviewer_assignments": [
            {
                "id": ra.id,
                "reviewer_id": ra.reviewer_id,
                "reviewer_name": f"{ra.reviewer.firstname} {ra.reviewer.lastname}" if ra.reviewer else None,
                "reviewer_email": ra.reviewer.email if ra.reviewer else None,
                "assigned_at": ra.assigned_at,
                "completed": ra.completed,
                "review": {
                    "id": ra.review.id,
                    "relevance_score": ra.review.relevance_score,
                    "methodology_score": ra.review.methodology_score,
                    "originality_score": ra.review.originality_score,
                    "overall_score": ra.review.overall_score,
                    "recommendation": ra.review.recommendation.value,
                    "comments": ra.review.comments,
                    "submitted_at": ra.review.submitted_at,
                } if ra.review else None,
            }
            for ra in a.reviewer_assignments
        ],
    }


@router.post("/", status_code=status.HTTP_201_CREATED)
def submit_abstract(
    request: Request,
    schema: AbstractSubmitSchema,
    current_user: Optional[dict] = Depends(get_optional_current_user),
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
):
    word_count = len(schema.abstract_text.split())

    # Resolve track text from track_id if provided
    track_text = schema.track
    track_id   = schema.track_id
    if track_id:
        track_record = db.query(EventTrack).filter(EventTrack.id == track_id).first()
        if track_record:
            track_text = f"{track_record.code}: {track_record.title}"
        else:
            track_id = None

    # Determine the submitter: logged-in user takes priority.
    # For anonymous submissions, link to an existing account if the email matches.
    # Account creation is deferred to the conference registration step.
    submitter_id = current_user["user_id"] if current_user else None
    submitter_email = current_user["username"] if current_user else None

    if not submitter_id and schema.authors:
        first_author = schema.authors[0]
        if first_author.email:
            existing = db.query(User).filter(
                User.email == first_author.email,
                User.deleted_at == None,
            ).first()
            if existing:
                submitter_id = existing.id
                submitter_email = existing.email

    abstract = Abstract(
        event_id=schema.event_id,
        submitted_by=submitter_id,
        title=schema.title,
        abstract_text=schema.abstract_text,
        keywords=schema.keywords,
        track=track_text,
        track_id=track_id,
        presentation_type=schema.presentation_type,
        word_count=word_count,
    )
    db.add(abstract)
    db.flush()

    for i, author in enumerate(schema.authors):
        db.add(AbstractAuthor(
            abstract_id=abstract.id,
            firstname=author.firstname,
            lastname=author.lastname,
            email=author.email,
            affiliation=author.affiliation,
            country=author.country,
            is_presenting=author.is_presenting,
            author_order=i,
        ))

    db.commit()
    db.refresh(abstract)
    abstract = db.query(Abstract).options(
        joinedload(Abstract.authors),
        joinedload(Abstract.submitter),
        joinedload(Abstract.event),
        joinedload(Abstract.event_track),
        joinedload(Abstract.reviewer_assignments),
    ).filter(Abstract.id == abstract.id).first()

    if submitter_id and submitter_email:
        dependency.log_activity(
            submitter_id,
            "SUBMIT_ABSTRACT",
            submitter_email,
            dependency.request_ip(request),
            f"Submitted abstract: {schema.title}",
        )

    first_author = schema.authors[0] if schema.authors else None
    return {
        "abstract": _serialize_abstract(abstract),
        "abstract_id": abstract.id,
        "author_firstname": first_author.firstname if first_author else None,
        "author_lastname": first_author.lastname if first_author else None,
        "author_email": first_author.email if first_author else None,
    }


@router.get("/")
def list_abstracts(
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
    event_id: int = Query(None),
    status_filter: str = Query(None, alias="status"),
    search: str = Query(None),
    track_id: int = Query(None),
    presentation_type: str = Query(None),
    skip: int = 0,
    limit: int = 50,
):
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])

    # Base filter query (no eager loads — used for count)
    q = db.query(Abstract).filter(Abstract.deleted_at == None)
    if event_id:
        q = q.filter(Abstract.event_id == event_id)
    if status_filter:
        q = q.filter(Abstract.status == status_filter)
    if track_id:
        q = q.filter(Abstract.track_id == track_id)
    if presentation_type:
        q = q.filter(Abstract.presentation_type == presentation_type)
    if search:
        q = q.filter(Abstract.title.ilike(f"%{search}%"))

    total = q.count()

    rows = (
        q.options(
            joinedload(Abstract.authors),
            joinedload(Abstract.submitter),
            joinedload(Abstract.event),
            joinedload(Abstract.event_track),
            joinedload(Abstract.reviewer_assignments).joinedload(AbstractReviewer.reviewer),
            joinedload(Abstract.reviewer_assignments).joinedload(AbstractReviewer.review),
        )
        .order_by(Abstract.created_at.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )

    return {
        "data": [_serialize_abstract(a) for a in rows],
        "total": total,
        "skip": skip,
        "limit": limit,
    }


@router.get("/export")
def export_abstracts(
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
    event_id: int = None,
    status_filter: str = Query(None, alias="status"),
    search: str = Query(None),
):
    auth_dependency.secure_access("EXPORT_ABSTRACTS", current_user["user_id"])
    from sqlalchemy import or_
    q = db.query(Abstract).options(
        joinedload(Abstract.authors),
        joinedload(Abstract.submitter),
        joinedload(Abstract.event),
        joinedload(Abstract.reviewer_assignments).joinedload(AbstractReviewer.reviewer),
        joinedload(Abstract.reviewer_assignments).joinedload(AbstractReviewer.review),
    ).filter(Abstract.deleted_at == None)
    if event_id:
        q = q.filter(Abstract.event_id == event_id)
    if status_filter:
        q = q.filter(Abstract.status == status_filter)
    if search:
        term = f"%{search}%"
        q = q.filter(or_(
            Abstract.title.ilike(term),
            Abstract.keywords.ilike(term),
        ))
    abstracts = q.order_by(Abstract.created_at.desc()).all()

    wb = Workbook()

    # ── Sheet 1: Abstracts ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Abstracts"

    header_fill = PatternFill("solid", start_color="0095B6")
    alt_fill    = PatternFill("solid", start_color="E8F4F8")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    body_font   = Font(name="Arial", size=10)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    headers = [
        "#", "Title", "Abstract Text", "Event", "Track", "Presentation Type", "Status",
        "Word Count", "Keywords", "Submitter Name", "Submitter Email",
        "First Author", "First Author Email", "First Author Affiliation",
        "First Author Country", "All Authors", "Reviewers Assigned",
        "Reviews Completed", "Avg Relevance", "Avg Methodology",
        "Avg Originality", "Avg Overall", "Date Submitted",
    ]

    ws.row_dimensions[1].height = 22
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    ws.freeze_panes = "A2"

    for ri, a in enumerate(abstracts, 2):
        ws.row_dimensions[ri].height = 16
        use_fill = alt_fill if ri % 2 == 0 else PatternFill("solid", start_color="FFFFFF")

        authors = sorted(a.authors, key=lambda x: x.author_order)
        first   = authors[0] if authors else None
        all_authors = "; ".join(
            f"{au.firstname} {au.lastname}" + (f" ({au.affiliation})" if au.affiliation else "")
            for au in authors
        )

        assignments = a.reviewer_assignments
        reviews = [r.review for r in assignments if r.review]
        avg = lambda field: round(sum(getattr(r, field) for r in reviews) / len(reviews), 1) if reviews else ""

        row = [
            a.id,
            a.title,
            a.abstract_text or "",
            a.event.event if a.event else "",
            a.track or "",
            a.presentation_type.value if a.presentation_type else "",
            a.status.value.replace("_", " ").title() if a.status else "",
            a.word_count or 0,
            a.keywords or "",
            f"{a.submitter.firstname} {a.submitter.lastname}" if a.submitter else "",
            a.submitter.email if a.submitter else "",
            f"{first.firstname} {first.lastname}" if first else "",
            first.email or "" if first else "",
            first.affiliation or "" if first else "",
            first.country or "" if first else "",
            all_authors,
            len(assignments),
            len(reviews),
            avg("relevance_score"),
            avg("methodology_score"),
            avg("originality_score"),
            avg("overall_score"),
            a.created_at.strftime("%d %b %Y") if a.created_at else "",
        ]

        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            cell.font = body_font
            cell.fill = use_fill
            cell.alignment = left

    # Auto column widths
    col_widths = [6, 40, 60, 30, 20, 16, 16, 10, 30, 22, 28, 22, 28, 28, 16, 50, 10, 10, 12, 12, 12, 12, 14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Authors ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Authors")
    ws2.sheet_properties.tabColor = "0095B6"
    auth_headers = ["Abstract #", "Abstract Title", "Author Order", "First Name", "Last Name",
                    "Email", "Affiliation", "Country", "Is Presenting"]
    ws2.row_dimensions[1].height = 22
    for ci, h in enumerate(auth_headers, 1):
        cell = ws2.cell(1, ci, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws2.freeze_panes = "A2"
    ri2 = 2
    for a in abstracts:
        for au in sorted(a.authors, key=lambda x: x.author_order):
            use_fill = alt_fill if ri2 % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
            row = [a.id, a.title, au.author_order + 1, au.firstname, au.lastname,
                   au.email or "", au.affiliation or "", au.country or "",
                   "Yes" if au.is_presenting else "No"]
            for ci, val in enumerate(row, 1):
                cell = ws2.cell(ri2, ci, val)
                cell.font = body_font
                cell.fill = use_fill
                cell.alignment = left
            ri2 += 1
    for ci, w in enumerate([10, 40, 12, 16, 16, 28, 28, 16, 12], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 3: Reviews ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Reviews")
    ws3.sheet_properties.tabColor = "0095B6"
    rev_headers = ["Abstract #", "Abstract Title", "Reviewer Name", "Reviewer Email",
                   "Completed", "Relevance", "Methodology", "Originality",
                   "Overall", "Recommendation", "Comments", "Review Date"]
    ws3.row_dimensions[1].height = 22
    for ci, h in enumerate(rev_headers, 1):
        cell = ws3.cell(1, ci, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws3.freeze_panes = "A2"
    ri3 = 2
    for a in abstracts:
        for ra in a.reviewer_assignments:
            use_fill = alt_fill if ri3 % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
            rv = ra.review
            row = [
                a.id, a.title,
                f"{ra.reviewer.firstname} {ra.reviewer.lastname}" if ra.reviewer else "",
                ra.reviewer.email if ra.reviewer else "",
                "Yes" if ra.completed else "No",
                rv.relevance_score if rv else "",
                rv.methodology_score if rv else "",
                rv.originality_score if rv else "",
                rv.overall_score if rv else "",
                rv.recommendation.value.replace("_", " ").title() if rv else "",
                rv.comments if rv else "",
                rv.submitted_at.strftime("%d %b %Y") if rv and rv.submitted_at else "",
            ]
            for ci, val in enumerate(row, 1):
                cell = ws3.cell(ri3, ci, val)
                cell.font = body_font
                cell.fill = use_fill
                cell.alignment = left
            ri3 += 1
    for ci, w in enumerate([10, 40, 22, 28, 10, 10, 12, 12, 10, 18, 50, 12], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    # Style sheet tabs
    ws.sheet_properties.tabColor = "0095B6"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "abstracts_export.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/pdf")
def export_abstracts_pdf(
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
    event_id: int = None,
    status_filter: str = Query(None, alias="status"),
):
    auth_dependency.secure_access("EXPORT_ABSTRACTS", current_user["user_id"])

    import unicodedata, re
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.colors import Color
    from reportlab.platypus import (
        BaseDocTemplate, PageTemplate, Frame,
        Paragraph, Spacer, HRFlowable, PageBreak,
        NextPageTemplate, KeepTogether, Table, TableStyle,
    )
    from reportlab.lib.enums import TA_JUSTIFY
    from collections import defaultdict
    from sqlalchemy import case as sa_case

    # ── Fetch data ────────────────────────────────────────────────────────────
    q = db.query(Abstract).options(
        joinedload(Abstract.authors),
        joinedload(Abstract.submitter),
        joinedload(Abstract.event),
        joinedload(Abstract.event_track),
    ).filter(Abstract.deleted_at == None)
    if event_id:
        q = q.filter(Abstract.event_id == event_id)
    if status_filter:
        q = q.filter(Abstract.status == status_filter)
    abstracts = q.order_by(
        sa_case((Abstract.track_id == None, 1), else_=0),
        Abstract.track_id,
        Abstract.title,
    ).all()

    event_name = abstracts[0].event.event if abstracts and abstracts[0].event else "ECSA Events"
    event_short = event_name if len(event_name) <= 75 else event_name[:72] + "..."

    groups = defaultdict(list)
    for a in abstracts:
        groups[(a.track_id, a.track or "No Track Assigned")].append(a)
    sorted_groups = sorted(groups.items(), key=lambda kv: (kv[0][0] is None, kv[0][1].lower()))

    # ── Colors ────────────────────────────────────────────────────────────────
    NAVY   = colors.HexColor("#1B3F6E")
    BLUE   = colors.HexColor("#1565C0")
    TEAL   = colors.HexColor("#0095B6")
    ORANGE = colors.HexColor("#F7941D")
    DGRAY  = colors.HexColor("#374151")
    LGRAY  = colors.HexColor("#9CA3AF")
    MGRAY  = colors.HexColor("#6B7280")
    WHITE  = colors.white

    # ── Layout ────────────────────────────────────────────────────────────────
    W, H = A4
    LM, RM = 1.6 * cm, 1.6 * cm
    TM, BM = 2.0 * cm, 2.0 * cm
    content_w = W - LM - RM
    content_h = H - TM - BM
    col_gap = 0.45 * cm
    col_w = (content_w - col_gap) / 2
    eff_w = col_w - 3          # effective width inside each column frame

    def xesc(s):
        return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def _wrap(text, max_chars):
        words, lines, curr = text.split(), [], ""
        for w in words:
            if len(curr) + len(w) + 1 <= max_chars:
                curr = (curr + " " + w).strip()
            else:
                if curr:
                    lines.append(curr)
                curr = w
        if curr:
            lines.append(curr)
        return lines

    # ── Page callbacks ────────────────────────────────────────────────────────
    def on_cover(c, doc):
        c.saveState()
        # Full navy background
        c.setFillColor(NAVY)
        c.rect(0, 0, W, H, fill=1, stroke=0)
        # Top strip
        c.setFillColor(Color(0.14, 0.26, 0.48))
        c.rect(0, H - 3.8 * cm, W, 3.8 * cm, fill=1, stroke=0)
        # Organisation name
        c.setFont("Helvetica-Bold", 11)
        c.setFillColor(WHITE)
        c.drawCentredString(W / 2, H - 1.7 * cm, "EAST, CENTRAL AND SOUTHERN AFRICA HEALTH COMMUNITY")
        c.setFont("Helvetica", 9)
        c.setFillColor(Color(0.56, 0.79, 0.95))
        c.drawCentredString(W / 2, H - 2.6 * cm, "Fostering Regional Cooperation for Better Health")
        # Teal accent line
        c.setStrokeColor(TEAL)
        c.setLineWidth(2)
        c.line(LM, H - 4.2 * cm, W - RM, H - 4.2 * cm)
        # Event name (wrapped)
        c.setFont("Helvetica-Bold", 17)
        c.setFillColor(WHITE)
        ev_lines = _wrap(event_name, 38)
        ey = H - 6.0 * cm
        for ln in ev_lines[:4]:
            c.drawCentredString(W / 2, ey, ln)
            ey -= 0.88 * cm
        # Orange bar
        c.setFillColor(ORANGE)
        c.rect(LM, ey - 0.2 * cm, content_w, 0.18 * cm, fill=1, stroke=0)
        # Blue "ABSTRACT BOOK" block
        ab_y = H * 0.27
        ab_h = H * 0.24
        c.setFillColor(BLUE)
        c.rect(0, ab_y, W * 0.70, ab_h, fill=1, stroke=0)
        c.setFont("Helvetica-Bold", 38)
        c.setFillColor(WHITE)
        c.drawString(LM, ab_y + ab_h * 0.60, "ABSTRACT")
        c.drawString(LM, ab_y + ab_h * 0.18, "BOOK")
        # Sub info
        c.setFont("Helvetica", 9.5)
        c.setFillColor(Color(0.56, 0.79, 0.95))
        n = len(abstracts)
        c.drawString(LM, ab_y - 0.75 * cm, f"{n} Abstract{'s' if n != 1 else ''}")
        c.drawString(LM, ab_y - 1.35 * cm, f"Generated {datetime.utcnow().strftime('%d %B %Y')}")
        # Bottom dark bar
        c.setFillColor(Color(0.05, 0.17, 0.37))
        c.rect(0, 0, W, 1.0 * cm, fill=1, stroke=0)
        c.restoreState()

    def on_divider(c, doc):
        c.saveState()
        c.setFillColor(BLUE)
        c.rect(0, 0, W, H, fill=1, stroke=0)
        c.setFillColor(Color(0.10, 0.46, 0.82))
        c.rect(0, H * 0.52, W, H * 0.48, fill=1, stroke=0)
        # Subtle diagonal wave lines
        c.setStrokeColor(Color(0.20, 0.60, 0.90, alpha=0.20))
        c.setLineWidth(50)
        c.line(-W * 0.3, H * 0.20, W * 1.4, H * 0.62)
        c.setLineWidth(30)
        c.line(-W * 0.2, H * 0.10, W * 1.3, H * 0.50)
        # Text
        c.setFont("Helvetica-Bold", 34)
        c.setFillColor(Color(0.80, 0.90, 0.96))
        c.drawString(2 * cm, H * 0.37, "ABSTRACTS")
        c.restoreState()

    def on_content(c, doc):
        c.saveState()
        fy = BM * 0.50
        c.setStrokeColor(TEAL)
        c.setLineWidth(0.5)
        c.line(LM, fy + 0.32 * cm, W - RM, fy + 0.32 * cm)
        c.setFont("Helvetica", 7)
        c.setFillColor(MGRAY)
        c.drawString(LM, fy, event_short)
        c.setFont("Helvetica-Bold", 8)
        c.setFillColor(TEAL)
        c.drawRightString(W - RM, fy, str(doc.page))
        c.restoreState()

    # ── Frames & document ─────────────────────────────────────────────────────
    cover_frame = Frame(LM, BM, content_w, content_h,
                        leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0, id='c')
    lf = Frame(LM, BM, col_w, content_h,
               leftPadding=0, rightPadding=3, topPadding=0, bottomPadding=0, id='l')
    rf = Frame(LM + col_w + col_gap, BM, col_w, content_h,
               leftPadding=3, rightPadding=0, topPadding=0, bottomPadding=0, id='r')

    buf = io.BytesIO()
    doc_obj = BaseDocTemplate(buf, pagesize=A4)
    doc_obj.addPageTemplates([
        PageTemplate('cover',   frames=[cover_frame], onPage=on_cover),
        PageTemplate('divider', frames=[cover_frame], onPage=on_divider),
        PageTemplate('main',    frames=[lf, rf],      onPage=on_content),
    ])

    # ── Styles ────────────────────────────────────────────────────────────────
    s_track_lbl = ParagraphStyle("TL", fontSize=8.5, textColor=WHITE,
                                 fontName="Helvetica-Bold", leading=12)
    s_title = ParagraphStyle("TI", fontSize=10, textColor=NAVY,
                             fontName="Helvetica-Bold", leading=14,
                             spaceBefore=8, spaceAfter=2)
    s_authors = ParagraphStyle("AU", fontSize=8, textColor=TEAL,
                               fontName="Helvetica-Oblique", leading=12, spaceAfter=2)
    s_meta = ParagraphStyle("ME", fontSize=7.5, textColor=LGRAY,
                             fontName="Helvetica", leading=11, spaceAfter=3)
    s_body = ParagraphStyle("BO", fontSize=8.5, textColor=DGRAY,
                             fontName="Helvetica", leading=13,
                             spaceAfter=3, alignment=TA_JUSTIFY)
    s_kw = ParagraphStyle("KW", fontSize=7.5, textColor=LGRAY,
                           fontName="Helvetica-Oblique", leading=11, spaceAfter=5)

    # ── Story ─────────────────────────────────────────────────────────────────
    story = []

    # Cover: on_cover draws everything; Spacer fills the frame
    story.append(Spacer(1, content_h))

    # Blue "ABSTRACTS" divider page
    story.append(NextPageTemplate('divider'))
    story.append(PageBreak())
    story.append(Spacer(1, content_h))

    # Two-column abstract pages
    story.append(NextPageTemplate('main'))
    story.append(PageBreak())

    for (track_id, track_label), track_abs in sorted_groups:
        # Teal track header box (fits within one column)
        th = Table([[Paragraph(xesc(track_label).upper(), s_track_lbl)]],
                   colWidths=[eff_w])
        th.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), TEAL),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 7),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        ]))
        story.append(th)
        story.append(Spacer(1, 0.15 * cm))

        for a in track_abs:
            authors_sorted = sorted(a.authors, key=lambda x: x.author_order)
            author_str = " | ".join(
                f"{au.firstname} {au.lastname}"
                + (f", {au.affiliation}" if au.affiliation else "")
                for au in authors_sorted
            )
            meta_parts = []
            if a.presentation_type:
                meta_parts.append(a.presentation_type.value.title())
            if a.word_count:
                meta_parts.append(f"{a.word_count} words")

            # Title + authors kept together (no orphaned title)
            header_block = [Paragraph(xesc(a.title), s_title)]
            if author_str:
                header_block.append(Paragraph(f"<i>by {xesc(author_str)}</i>", s_authors))
            if meta_parts:
                header_block.append(Paragraph(xesc(" · ".join(meta_parts)), s_meta))
            story.append(KeepTogether(header_block))

            if a.abstract_text:
                safe = xesc(a.abstract_text).replace("\n\n", "<br/><br/>").replace("\n", " ")
                story.append(Paragraph(safe, s_body))
            if a.keywords:
                story.append(Paragraph(f"<i>Keywords: {xesc(a.keywords)}</i>", s_kw))
            story.append(HRFlowable(width="100%", thickness=0.4,
                                    color=colors.HexColor("#D1D5DB"), spaceAfter=5))

        story.append(Spacer(1, 0.3 * cm))

    doc_obj.build(story)
    buf.seek(0)

    safe_e = unicodedata.normalize("NFKD", event_name).encode("ascii", "ignore").decode("ascii")
    safe_e = re.sub(r"[^\w\s-]", "", safe_e).strip().replace(" ", "_")[:40] or "abstracts"
    filename = f"book_of_abstracts_{safe_e}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/my-submissions")
def my_submissions(
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    abstracts = db.query(Abstract).options(
        joinedload(Abstract.authors),
        joinedload(Abstract.event),
        joinedload(Abstract.reviewer_assignments),
    ).filter(
        Abstract.submitted_by == current_user["user_id"],
        Abstract.deleted_at == None,
    ).order_by(Abstract.created_at.desc()).all()
    return [_serialize_abstract(a) for a in abstracts]


@router.get("/my-reviews")
def my_reviews(
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    assignments = db.query(AbstractReviewer).options(
        joinedload(AbstractReviewer.abstract).joinedload(Abstract.authors),
        joinedload(AbstractReviewer.abstract).joinedload(Abstract.event),
        joinedload(AbstractReviewer.review),
    ).filter(
        AbstractReviewer.reviewer_id == current_user["user_id"],
    ).all()
    return [
        {
            "assignment_id": a.id,
            "assigned_at": a.assigned_at,
            "completed": a.completed,
            "abstract": _serialize_abstract(a.abstract),
            "review": {
                "id": a.review.id,
                "relevance_score": a.review.relevance_score,
                "methodology_score": a.review.methodology_score,
                "originality_score": a.review.originality_score,
                "overall_score": a.review.overall_score,
                "recommendation": a.review.recommendation.value,
                "comments": a.review.comments,
                "submitted_at": a.review.submitted_at,
            } if a.review else None,
        }
        for a in assignments
    ]


@router.get("/reviewers/candidates")
def list_reviewer_candidates(
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
):
    """Return all users marked as reviewers (is_reviewer=True) for the assignment picker."""
    auth_dependency.secure_access("MANAGE_REVIEWERS", current_user["user_id"])
    users = (
        db.query(User)
        .filter(User.is_reviewer == True, User.deleted_at == None)
        .order_by(User.firstname, User.lastname)
        .all()
    )
    return [
        {"id": u.id, "firstname": u.firstname, "lastname": u.lastname, "email": u.email}
        for u in users
    ]


@router.get("/reviewers")
def list_reviewers(
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
    event_id: int = None,
):
    auth_dependency.secure_access("MANAGE_REVIEWERS", current_user["user_id"])
    q = db.query(AbstractReviewer).options(
        joinedload(AbstractReviewer.reviewer),
        joinedload(AbstractReviewer.abstract).joinedload(Abstract.event),
        joinedload(AbstractReviewer.review),
    )
    if event_id:
        from sqlalchemy import join
        q = q.join(Abstract, AbstractReviewer.abstract_id == Abstract.id).filter(Abstract.event_id == event_id)
    assignments = q.all()

    reviewers: dict = {}
    for a in assignments:
        rid = a.reviewer_id
        if rid not in reviewers:
            reviewers[rid] = {
                "reviewer_id": rid,
                "name": f"{a.reviewer.firstname} {a.reviewer.lastname}" if a.reviewer else "",
                "email": a.reviewer.email if a.reviewer else "",
                "total": 0,
                "completed": 0,
                "assignments": [],
            }
        reviewers[rid]["total"] += 1
        if a.completed:
            reviewers[rid]["completed"] += 1
        reviewers[rid]["assignments"].append({
            "assignment_id": a.id,
            "abstract_id": a.abstract_id,
            "abstract_title": a.abstract.title if a.abstract else "",
            "event": a.abstract.event.event if a.abstract and a.abstract.event else "",
            "event_id": a.abstract.event_id if a.abstract else None,
            "assigned_at": a.assigned_at,
            "completed": a.completed,
        })
    return list(reviewers.values())


@router.get("/stats")
def abstract_stats(
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
    event_id: int = Query(None),
):
    """Aggregated stats for the abstract dashboard."""
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])

    from sqlalchemy import func as sqlfunc, case as sa_case
    from models.models import AbstractAuthor, Event

    base = db.query(Abstract).filter(Abstract.deleted_at == None)
    if event_id:
        base = base.filter(Abstract.event_id == event_id)

    by_status_rows = (
        base.with_entities(Abstract.status, sqlfunc.count().label("n"))
        .group_by(Abstract.status).all()
    )
    by_status = [{"status": r.status.value if r.status else "unknown", "count": r.n}
                 for r in by_status_rows]

    by_type_rows = (
        base.with_entities(Abstract.presentation_type, sqlfunc.count().label("n"))
        .group_by(Abstract.presentation_type).all()
    )
    by_type = [{"type": r.presentation_type.value if r.presentation_type else "unspecified", "count": r.n}
               for r in by_type_rows]

    by_track_rows = (
        base.with_entities(
            Abstract.track, Abstract.track_id,
            sqlfunc.count().label("total"),
            sqlfunc.sum(sa_case((Abstract.status == AbstractStatus.accepted, 1), else_=0)).label("accepted"),
        )
        .group_by(Abstract.track, Abstract.track_id)
        .order_by(sqlfunc.count().desc()).all()
    )
    by_track = [
        {"track": r.track or "Untracked", "track_id": r.track_id,
         "total": r.total, "accepted": int(r.accepted or 0)}
        for r in by_track_rows
    ]

    country_total_q = (
        db.query(AbstractAuthor.country, sqlfunc.count(sqlfunc.distinct(AbstractAuthor.abstract_id)).label("total"))
        .join(Abstract, Abstract.id == AbstractAuthor.abstract_id)
        .filter(Abstract.deleted_at == None, AbstractAuthor.country != None, AbstractAuthor.country != "")
    )
    if event_id:
        country_total_q = country_total_q.filter(Abstract.event_id == event_id)
    country_total_rows = (
        country_total_q.group_by(AbstractAuthor.country)
        .order_by(sqlfunc.count(sqlfunc.distinct(AbstractAuthor.abstract_id)).desc()).all()
    )

    accepted_id_q = db.query(Abstract.id).filter(
        Abstract.status == AbstractStatus.accepted,
        Abstract.deleted_at == None,
        *([Abstract.event_id == event_id] if event_id else [])
    )
    country_accepted_rows = (
        db.query(AbstractAuthor.country, sqlfunc.count(sqlfunc.distinct(AbstractAuthor.abstract_id)).label("accepted"))
        .filter(AbstractAuthor.abstract_id.in_(accepted_id_q),
                AbstractAuthor.country != None, AbstractAuthor.country != "")
        .group_by(AbstractAuthor.country).all()
    )
    def _norm_country(c):
        return (c or "").strip().rstrip(".,;").strip()

    # Merge accepted counts using normalised country name as key
    accepted_by_norm = {}
    for r in country_accepted_rows:
        key = _norm_country(r.country)
        accepted_by_norm[key] = accepted_by_norm.get(key, 0) + r.accepted

    # Merge totals using normalised country name; pick the most-common spelling for display
    totals_by_norm: dict = {}
    for r in country_total_rows:
        key = _norm_country(r.country)
        if key not in totals_by_norm:
            totals_by_norm[key] = {"country": r.country, "total": 0}
        totals_by_norm[key]["total"] += r.total

    by_country = [
        {"country": v["country"], "total": v["total"], "accepted": accepted_by_norm.get(k, 0)}
        for k, v in sorted(totals_by_norm.items(), key=lambda x: -x[1]["total"])
    ]

    esw_names = {"eswatini", "swaziland"}
    eswatini_accepted = sum(
        c["accepted"] for c in by_country
        if (c["country"] or "").strip().lower() in esw_names
    )

    # Abstracts where reviewer consensus = reject
    # Rule: rejected if reject_count >= 2, OR rejected by all reviewers when < 2 assigned
    from models.models import AbstractReviewer, AbstractReview, ReviewRecommendation
    rev_q = (
        db.query(
            AbstractReviewer.abstract_id,
            sqlfunc.count(AbstractReviewer.id).label("total_reviewers"),
            sqlfunc.sum(
                sa_case((AbstractReview.recommendation == ReviewRecommendation.reject, 1), else_=0)
            ).label("reject_count"),
        )
        .join(AbstractReview, AbstractReview.assignment_id == AbstractReviewer.id, isouter=True)
        .join(Abstract, Abstract.id == AbstractReviewer.abstract_id)
        .filter(Abstract.deleted_at == None)
    )
    if event_id:
        rev_q = rev_q.filter(Abstract.event_id == event_id)
    rev_rows = rev_q.group_by(AbstractReviewer.abstract_id).all()

    reviewer_rejected = sum(
        1 for r in rev_rows
        if r.reject_count and r.total_reviewers and (
            r.reject_count >= 2 or r.reject_count == r.total_reviewers
        )
    )

    events_list = [
        {"id": e.id, "event": e.event}
        for e in db.query(Event).filter(Event.deleted_at == None).order_by(Event.id.desc()).all()
    ]
    total_all = sum(r["count"] for r in by_status)
    total_accepted = next((r["count"] for r in by_status if r["status"] == "accepted"), 0)
    return {
        "total": total_all, "total_accepted": total_accepted,
        "eswatini_accepted": eswatini_accepted,
        "reviewer_rejected": reviewer_rejected,
        "by_status": by_status, "by_type": by_type,
        "by_track": by_track, "by_country": by_country,
        "events": events_list,
    }


@router.get("/{abstract_id}")
def get_abstract(
    abstract_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    abstract = db.query(Abstract).options(
        joinedload(Abstract.authors),
        joinedload(Abstract.submitter),
        joinedload(Abstract.event),
        joinedload(Abstract.event_track),
        joinedload(Abstract.reviewer_assignments).joinedload(AbstractReviewer.reviewer),
        joinedload(Abstract.reviewer_assignments).joinedload(AbstractReviewer.review),
    ).filter(Abstract.id == abstract_id, Abstract.deleted_at == None).first()
    if not abstract:
        raise HTTPException(status_code=404, detail="Abstract not found")
    return _serialize_abstract(abstract)


@router.put("/{abstract_id}")
def update_abstract(
    abstract_id: int,
    schema: AbstractUpdateSchema,
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    abstract = db.query(Abstract).filter(Abstract.id == abstract_id, Abstract.deleted_at == None).first()
    if not abstract:
        raise HTTPException(status_code=404, detail="Abstract not found")

    if schema.title: abstract.title = schema.title
    if schema.abstract_text:
        abstract.abstract_text = schema.abstract_text
        abstract.word_count = len(schema.abstract_text.split())
    if schema.keywords is not None: abstract.keywords = schema.keywords
    if schema.track is not None: abstract.track = schema.track
    if schema.track_id is not None: abstract.track_id = schema.track_id
    if schema.presentation_type: abstract.presentation_type = schema.presentation_type
    if schema.status: abstract.status = schema.status

    if schema.authors is not None:
        db.query(AbstractAuthor).filter(AbstractAuthor.abstract_id == abstract_id).delete()
        for i, author in enumerate(schema.authors):
            db.add(AbstractAuthor(
                abstract_id=abstract_id,
                firstname=author.firstname,
                lastname=author.lastname,
                email=author.email,
                affiliation=author.affiliation,
                country=author.country,
                is_presenting=author.is_presenting,
                author_order=i,
            ))

    db.commit()
    db.refresh(abstract)
    abstract = db.query(Abstract).options(
        joinedload(Abstract.authors),
        joinedload(Abstract.submitter),
        joinedload(Abstract.event),
        joinedload(Abstract.reviewer_assignments).joinedload(AbstractReviewer.reviewer),
        joinedload(Abstract.reviewer_assignments).joinedload(AbstractReviewer.review),
    ).filter(Abstract.id == abstract_id).first()
    return _serialize_abstract(abstract)


@router.put("/{abstract_id}/status")
def update_status(
    abstract_id: int,
    payload: dict,
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    abstract = db.query(Abstract).filter(Abstract.id == abstract_id, Abstract.deleted_at == None).first()
    if not abstract:
        raise HTTPException(status_code=404, detail="Abstract not found")
    abstract.status = payload.get("status", abstract.status)
    db.commit()
    return {"message": "Status updated"}


@router.post("/bulk-assign-reviewer", status_code=status.HTTP_201_CREATED)
def bulk_assign_reviewer(
    schema: BulkAssignReviewerSchema,
    background_tasks: BackgroundTasks,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
):
    """Assign one reviewer to multiple abstracts and send a single consolidated email."""
    auth_dependency.secure_access("MANAGE_REVIEWERS", current_user["user_id"])

    reviewer = db.query(User).filter(User.id == schema.reviewer_id).first()
    if not reviewer:
        raise HTTPException(status_code=404, detail="Reviewer not found")

    assigner = db.query(User).filter(User.id == current_user["user_id"]).first()
    assigner_name  = f"{assigner.firstname} {assigner.lastname}" if assigner else "ECSA Secretariat"
    assigner_email = assigner.email if assigner else None

    assigned_abstracts = []
    skipped = []

    for abstract_id in schema.abstract_ids:
        abstract = db.query(Abstract).options(joinedload(Abstract.event)).filter(
            Abstract.id == abstract_id, Abstract.deleted_at == None
        ).first()
        if not abstract:
            skipped.append({"id": abstract_id, "reason": "not found"})
            continue
        existing = db.query(AbstractReviewer).filter(
            AbstractReviewer.abstract_id == abstract_id,
            AbstractReviewer.reviewer_id == schema.reviewer_id,
        ).first()
        if existing:
            skipped.append({"id": abstract_id, "reason": "already assigned"})
            continue
        db.add(AbstractReviewer(
            abstract_id=abstract_id,
            reviewer_id=schema.reviewer_id,
            assigned_by=current_user["user_id"],
        ))
        if abstract.status == AbstractStatus.submitted:
            abstract.status = AbstractStatus.under_review
        assigned_abstracts.append(abstract)

    if not assigned_abstracts:
        db.commit()
        return {"message": "No new assignments made", "assigned": 0, "skipped": skipped}

    import utils.mailer_util as mailer_util

    new_password = None
    if not reviewer.credentials_sent:
        new_password = auth_dependency.generate_random_password()
        reviewer.hashed_password = auth_dependency.hash_password(new_password)
        reviewer.credentials_sent = True

    db.commit()

    mailer_util.reviewer_bulk_assignment_email(
        recipient_email=reviewer.email,
        firstname=reviewer.firstname,
        password=new_password,
        abstracts=[
            {"title": a.title, "event": a.event.event if a.event else ""}
            for a in assigned_abstracts
        ],
        assigned_by_name=assigner_name,
        assigned_by_email=assigner_email,
        sent_by_user_id=current_user["user_id"],
        background_tasks=background_tasks,
        db=db,
    )

    return {
        "message": f"{len(assigned_abstracts)} abstract(s) assigned to {reviewer.firstname} {reviewer.lastname}",
        "assigned": len(assigned_abstracts),
        "skipped": skipped,
    }


@router.post("/{abstract_id}/assign-reviewer", status_code=status.HTTP_201_CREATED)
def assign_reviewer(
    abstract_id: int,
    schema: AssignReviewerSchema,
    background_tasks: BackgroundTasks,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
):
    auth_dependency.secure_access("MANAGE_REVIEWERS", current_user["user_id"])
    abstract = db.query(Abstract).options(
        joinedload(Abstract.event)
    ).filter(Abstract.id == abstract_id, Abstract.deleted_at == None).first()
    if not abstract:
        raise HTTPException(status_code=404, detail="Abstract not found")

    existing = db.query(AbstractReviewer).filter(
        AbstractReviewer.abstract_id == abstract_id,
        AbstractReviewer.reviewer_id == schema.reviewer_id,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Reviewer already assigned")

    reviewer = db.query(User).filter(User.id == schema.reviewer_id).first()
    if not reviewer:
        raise HTTPException(status_code=404, detail="Reviewer not found")

    assigner = db.query(User).filter(User.id == current_user["user_id"]).first()
    assigner_name  = f"{assigner.firstname} {assigner.lastname}" if assigner else "ECSA Secretariat"
    assigner_email = assigner.email if assigner else None

    assignment = AbstractReviewer(
        abstract_id=abstract_id,
        reviewer_id=schema.reviewer_id,
        assigned_by=current_user["user_id"],
    )
    db.add(assignment)

    if abstract.status == AbstractStatus.submitted:
        abstract.status = AbstractStatus.under_review

    import utils.mailer_util as mailer_util

    if not reviewer.credentials_sent:
        # First time this user is contacted — generate a fresh password and
        # include it in the assignment email so they can log in.
        new_password = auth_dependency.generate_random_password()
        reviewer.hashed_password = auth_dependency.hash_password(new_password)
        reviewer.credentials_sent = True
        db.commit()
        mailer_util.reviewer_assignment_email(
            recipient_email=reviewer.email,
            firstname=reviewer.firstname,
            password=new_password,
            abstract_title=abstract.title,
            event_name=abstract.event.event if abstract.event else None,
            assigned_by_name=assigner_name,
            assigned_by_email=assigner_email,
            sent_by_user_id=current_user["user_id"],
            background_tasks=background_tasks,
            db=db,
        )
    else:
        # Reviewer already has their credentials — just notify them of the new
        # assignment without touching their password.
        db.commit()
        mailer_util.reviewer_assignment_email(
            recipient_email=reviewer.email,
            firstname=reviewer.firstname,
            password=None,
            abstract_title=abstract.title,
            event_name=abstract.event.event if abstract.event else None,
            assigned_by_name=assigner_name,
            assigned_by_email=assigner_email,
            sent_by_user_id=current_user["user_id"],
            background_tasks=background_tasks,
            db=db,
        )

    return {"message": "Reviewer assigned", "reviewer_name": f"{reviewer.firstname} {reviewer.lastname}"}


@router.delete("/{abstract_id}/reviewers/{reviewer_id}")
def remove_reviewer(
    abstract_id: int,
    reviewer_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
):
    auth_dependency.secure_access("MANAGE_REVIEWERS", current_user["user_id"])
    assignment = db.query(AbstractReviewer).filter(
        AbstractReviewer.abstract_id == abstract_id,
        AbstractReviewer.reviewer_id == reviewer_id,
    ).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    db.delete(assignment)
    db.commit()
    return {"message": "Reviewer removed"}


@router.delete("/{abstract_id}")
def delete_abstract(
    abstract_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
):
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])
    abstract = db.query(Abstract).filter(Abstract.id == abstract_id, Abstract.deleted_at == None).first()
    if not abstract:
        raise HTTPException(status_code=404, detail="Abstract not found")
    abstract.deleted_at = datetime.utcnow()
    db.commit()
    return {"message": "Abstract deleted"}


@router.post("/reviewers/create", status_code=status.HTTP_201_CREATED)
def create_reviewer(
    schema: CreateReviewerSchema,
    current_user: user_dependency,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
):
    """Create a new user account designated as a reviewer and send them login credentials."""
    auth_dependency.secure_access("MANAGE_REVIEWERS", current_user["user_id"])

    existing = db.query(User).filter(User.email == schema.email).first()
    if existing:
        # User already exists — mark them as a reviewer and return their data
        existing.is_reviewer = True
        db.commit()
        return {
            "id": existing.id,
            "firstname": existing.firstname,
            "lastname": existing.lastname,
            "email": existing.email,
        }

    # Get or create the phone (phone optional for reviewers — use email as fallback stub)
    phone = schema.phone or schema.email

    password = auth_dependency.generate_random_password()
    hashed = auth_dependency.hash_password(password)

    user = User(
        firstname=schema.firstname,
        lastname=schema.lastname,
        email=schema.email,
        phone=phone,
        hashed_password=hashed,
        verified=True,
        is_reviewer=True,
    )
    db.add(user)
    db.flush()

    # Assign "User" role so they can log in
    user_role = db.query(Role).filter(Role.role == "User").first()
    if user_role:
        db.add(UserRole(user_id=user.id, role_id=user_role.id))

    db.commit()
    db.refresh(user)

    # Credentials are intentionally NOT sent here — they will be emailed
    # automatically when the reviewer is first assigned to an abstract.

    return {
        "id": user.id,
        "firstname": user.firstname,
        "lastname": user.lastname,
        "email": user.email,
    }


@router.post("/reviews/{assignment_id}", status_code=status.HTTP_201_CREATED)
def submit_review(
    assignment_id: int,
    schema: AbstractReviewSchema,
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    assignment = db.query(AbstractReviewer).filter(
        AbstractReviewer.id == assignment_id,
        AbstractReviewer.reviewer_id == current_user["user_id"],
    ).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    if assignment.review:
        raise HTTPException(status_code=400, detail="Review already submitted")

    review = AbstractReview(
        assignment_id=assignment_id,
        relevance_score=schema.relevance_score,
        methodology_score=schema.methodology_score,
        originality_score=schema.originality_score,
        overall_score=schema.overall_score,
        recommendation=schema.recommendation,
        comments=schema.comments,
        confidential_comments=schema.confidential_comments,
    )
    db.add(review)
    assignment.completed = True
    db.commit()
    return {"message": "Review submitted successfully"}


@router.put("/reviews/{assignment_id}")
def update_review(
    assignment_id: int,
    schema: AbstractReviewSchema,
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    assignment = db.query(AbstractReviewer).filter(
        AbstractReviewer.id == assignment_id,
        AbstractReviewer.reviewer_id == current_user["user_id"],
    ).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    if not assignment.review:
        raise HTTPException(status_code=400, detail="No review submitted yet")

    review = assignment.review
    review.relevance_score    = schema.relevance_score
    review.methodology_score  = schema.methodology_score
    review.originality_score  = schema.originality_score
    review.overall_score      = schema.overall_score
    review.recommendation     = schema.recommendation
    review.comments           = schema.comments
    review.confidential_comments = schema.confidential_comments
    db.commit()
    return {"message": "Review updated successfully"}


@router.get("/{abstract_id}/reviews")
def get_reviews(
    abstract_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    assignments = db.query(AbstractReviewer).options(
        joinedload(AbstractReviewer.reviewer),
        joinedload(AbstractReviewer.review),
    ).filter(AbstractReviewer.abstract_id == abstract_id).all()
    return [
        {
            "assignment_id": a.id,
            "reviewer": f"{a.reviewer.firstname} {a.reviewer.lastname}" if a.reviewer else None,
            "completed": a.completed,
            "review": {
                "relevance_score": a.review.relevance_score,
                "methodology_score": a.review.methodology_score,
                "originality_score": a.review.originality_score,
                "overall_score": a.review.overall_score,
                "recommendation": a.review.recommendation.value,
                "comments": a.review.comments,
                "submitted_at": a.review.submitted_at,
            } if a.review else None,
        }
        for a in assignments
    ]


# ══════════════════════════════════════════════════════════════════════════════
#  TRACKS  endpoints
# ══════════════════════════════════════════════════════════════════════════════

from pydantic import BaseModel as PydanticBase
from typing import Optional

class TrackUpdateSchema(PydanticBase):
    code:       Optional[str] = None
    title:      Optional[str] = None
    theme:      Optional[str] = None
    sort_order: Optional[int] = None


def _serialize_track(t: EventTrack, abstract_count: int = 0) -> dict:
    return {
        "id":             t.id,
        "event_id":       t.event_id,
        "event":          t.event.event if t.event else None,
        "code":           t.code,
        "title":          t.title,
        "theme":          t.theme,
        "sort_order":     t.sort_order,
        "abstract_count": abstract_count,
        "created_at":     t.created_at,
    }


@router.get("/tracks/list")
def list_tracks(
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
    event_id: int = None,
):
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])
    from sqlalchemy import func as sqlfunc
    q = db.query(
        EventTrack,
        sqlfunc.count(Abstract.id).label("abstract_count")
    ).outerjoin(
        Abstract,
        (Abstract.track_id == EventTrack.id) & (Abstract.deleted_at == None)
    ).options(joinedload(EventTrack.event))
    if event_id:
        q = q.filter(EventTrack.event_id == event_id)
    rows = q.group_by(EventTrack.id).order_by(EventTrack.sort_order, EventTrack.code).all()
    return [_serialize_track(t, cnt) for t, cnt in rows]


@router.get("/tracks/public")
def list_tracks_public(
    db: Session = Depends(get_db),
    event_id: int = None,
):
    """Public endpoint — no auth required. Returns tracks for the submission form."""
    q = db.query(EventTrack)
    if event_id:
        q = q.filter(EventTrack.event_id == event_id)
    rows = q.order_by(EventTrack.sort_order, EventTrack.code).all()
    return [
        {"id": t.id, "code": t.code, "title": t.title, "theme": t.theme, "sort_order": t.sort_order}
        for t in rows
    ]


@router.put("/tracks/{track_id}")
def update_track(
    track_id: int,
    schema: TrackUpdateSchema,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
):
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])
    track = db.query(EventTrack).options(joinedload(EventTrack.event)).filter(EventTrack.id == track_id).first()
    if not track:
        raise HTTPException(status_code=404, detail="Track not found")
    if schema.code       is not None: track.code       = schema.code
    if schema.title      is not None: track.title      = schema.title
    if schema.theme      is not None: track.theme      = schema.theme
    if schema.sort_order is not None: track.sort_order = schema.sort_order
    if schema.code is not None or schema.title is not None:
        new_text = f"{track.code}: {track.title}"
        db.query(Abstract).filter(Abstract.track_id == track_id).update(
            {"track": new_text}, synchronize_session=False
        )
    db.commit()
    db.refresh(track)
    from sqlalchemy import func as sqlfunc
    cnt = db.query(sqlfunc.count(Abstract.id)).filter(
        Abstract.track_id == track_id, Abstract.deleted_at == None
    ).scalar() or 0
    return _serialize_track(track, cnt)


# ══════════════════════════════════════════════════════════════════════════════
#  BULK ACCEPT + NOTIFY endpoints
# ══════════════════════════════════════════════════════════════════════════════

from pydantic import BaseModel as _PydanticBase
from typing import List as _List

class BulkAcceptItem(_PydanticBase):
    id: int
    presentation_type: Optional[str] = None   # "oral" | "poster"

class BulkAcceptSchema(_PydanticBase):
    items: _List[BulkAcceptItem]

class NotifyAcceptedSchema(_PydanticBase):
    abstract_ids: Optional[_List[int]] = None   # None → all accepted for event
    event_id: Optional[int] = None
    portal_url: str = "https://events.ecsahc.org"
    test_email: Optional[str] = None            # if set, all emails go here
    skip_notified: bool = True                  # skip abstracts already notified


@router.post("/bulk-accept")
def bulk_accept(
    schema: BulkAcceptSchema,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
):
    """Set status=accepted (and optionally presentation_type) for a list of abstracts."""
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])
    ids = [i.id for i in schema.items]
    abstracts = (
        db.query(Abstract)
        .filter(Abstract.id.in_(ids), Abstract.deleted_at == None)
        .all()
    )
    id_to_type = {i.id: i.presentation_type for i in schema.items}
    updated = []
    not_found = list(set(ids) - {a.id for a in abstracts})
    for a in abstracts:
        a.status = AbstractStatus.accepted
        if id_to_type.get(a.id):
            a.presentation_type = id_to_type[a.id]
        updated.append(a.id)
    db.commit()
    return {"updated": len(updated), "not_found": not_found}


@router.post("/notify-acceptance")
def notify_acceptance(
    schema: NotifyAcceptedSchema,
    background_tasks: BackgroundTasks,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dep),
):
    """Send acceptance notification emails to authors of accepted abstracts."""
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])

    from models.models import Event, AbstractAuthor

    q = db.query(Abstract).filter(
        Abstract.status == AbstractStatus.accepted,
        Abstract.deleted_at == None,
    )
    if schema.abstract_ids:
        q = q.filter(Abstract.id.in_(schema.abstract_ids))
    else:
        if schema.skip_notified:
            q = q.filter(Abstract.acceptance_notified_at == None)
        if schema.event_id:
            q = q.filter(Abstract.event_id == schema.event_id)

    abstracts = q.options(
        joinedload(Abstract.event),
        joinedload(Abstract.authors),
    ).all()

    if not abstracts:
        raise HTTPException(status_code=404, detail="No accepted abstracts found matching criteria")

    import utils.mailer_util as _mailer
    import smtplib, os, time
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from datetime import datetime, timezone

    _eswatini_names = {"eswatini", "swaziland"}

    # Build the full job list before touching SMTP
    jobs = []
    for a in abstracts:
        event_name = a.event.event if a.event else "ECSA BPF"
        p_type = a.presentation_type.value if a.presentation_type else "oral"
        is_eswatini = any(
            (au.country or "").strip().lower() in _eswatini_names
            for au in (a.authors or [])
        )
        recipients = []
        if a.authors:
            presenting = [au for au in a.authors if au.is_presenting and au.email]
            targets = presenting or [au for au in a.authors if au.email][:1]
            for au in targets:
                recipients.append((au.firstname or "Presenter", au.email))
        if not recipients:
            submitter = db.query(User).filter(User.id == a.submitted_by).first()
            if submitter and submitter.email:
                recipients.append((submitter.firstname or submitter.email.split("@")[0], submitter.email))
        for firstname, email in recipients:
            jobs.append({
                "abstract": a, "firstname": firstname,
                "email": schema.test_email or email,
                "real_email": email,
                "event_name": event_name, "p_type": p_type, "is_eswatini": is_eswatini,
            })

    sent = []
    failed = []

    def _send_all(jobs, portal_url, sent_by_user_id, test_email, db_session):
        smtp_host     = os.getenv("SMTP_HOST", "")
        smtp_port     = int(os.getenv("SMTP_PORT", "587"))
        smtp_username = os.getenv("SMTP_USERNAME", "")
        smtp_password = os.getenv("SMTP_PASSWORD", "")
        from_email    = os.getenv("SMTP_FROM_EMAIL", smtp_username)
        templates     = _mailer.templates

        try:
            server = smtplib.SMTP(smtp_host, smtp_port)
            server.starttls()
            server.login(smtp_username, smtp_password)
        except Exception as e:
            for j in jobs:
                failed.append({"abstract_id": j["abstract"].id, "email": j["email"], "error": str(e)})
            return

        for j in jobs:
            try:
                subject = f"Congratulations! Your Abstract Has Been Accepted – {j['event_name']}"
                tmpl = templates.get_template("abstract_acceptance_template.html")
                body = tmpl.render(
                    subject=subject, firstname=j["firstname"],
                    abstract_title=j["abstract"].title,
                    presentation_type=j["p_type"],
                    event_name=j["event_name"],
                    portal_url=portal_url,
                    ppt_template_url="",
                    is_eswatini=j["is_eswatini"],
                    year=_mailer.YEAR,
                )
                msg = MIMEMultipart()
                msg["From"]    = f"{_mailer.FROM_NAME} <{from_email}>"
                msg["To"]      = j["email"]
                msg["Subject"] = subject
                msg.attach(MIMEText(body, "html"))
                server.sendmail(smtp_username, j["email"], msg.as_string())
                sent.append({"abstract_id": j["abstract"].id, "email": j["email"]})
                log_id = _mailer._create_email_log(db_session, j["email"], subject,
                                                   "abstract_acceptance", sent_by_user_id, None, body)
                if log_id:
                    _mailer._update_email_log(db_session, log_id, "sent")
                if not test_email:
                    j["abstract"].acceptance_notified_at = datetime.now(timezone.utc)
                    db_session.add(j["abstract"])
                time.sleep(0.3)   # 300 ms gap — stay within Gmail's rate limit
            except smtplib.SMTPServerDisconnected:
                # Reconnect once on dropped connection
                try:
                    server = smtplib.SMTP(smtp_host, smtp_port)
                    server.starttls()
                    server.login(smtp_username, smtp_password)
                    server.sendmail(smtp_username, j["email"], msg.as_string())
                    sent.append({"abstract_id": j["abstract"].id, "email": j["email"]})
                except Exception as retry_e:
                    failed.append({"abstract_id": j["abstract"].id, "email": j["email"], "error": str(retry_e)})
            except Exception as exc:
                failed.append({"abstract_id": j["abstract"].id, "email": j["email"], "error": str(exc)})

        try:
            server.quit()
        except Exception:
            pass
        db_session.commit()

    background_tasks.add_task(
        _send_all, jobs, schema.portal_url,
        current_user["user_id"], schema.test_email, db,
    )

    return {
        "sent": len(jobs),   # optimistic count returned immediately
        "failed": 0,
        "details": [],
        "errors": failed,
    }

