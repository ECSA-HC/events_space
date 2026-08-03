import math, os, io, secrets
from pydantic import BaseModel
import uuid
import shutil
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from io import BytesIO
from typing import Literal
from fastapi.responses import JSONResponse
from sqlalchemy import or_
from fastapi import status, HTTPException, File, Form, UploadFile
from typing import Annotated
from core.database import get_db
from sqlalchemy.orm import Session, joinedload, selectinload
from datetime import datetime, timedelta
from dependencies.auth_dependency import Auth
from dependencies.dependency import Dependency
from dependencies.auth_dependency import get_current_user, get_optional_current_user
from typing import Optional
from fastapi import APIRouter, BackgroundTasks, HTTPException, Depends, Query, Request
from models.models import Event, User, Registration, Document, Link, Payment, ParticipationRole, UserProfile
from schemas.events_space import EventSchema, EventUpdateSchema, RegistrationSchema, LinkSchema, PaymentSubmitSchema
from PIL import Image
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
import qrcode
from fastapi.responses import StreamingResponse
from reportlab.pdfgen import canvas
import unicodedata
import re
import urllib.parse


router = APIRouter()

user_dependency = Annotated[dict, Depends(get_current_user)]


def get_dependency(db: Session = Depends(get_db)) -> Dependency:
    return Dependency(db)


def get_auth_dependency(db: Session = Depends(get_db)) -> Auth:
    return Auth(db)


CLIENT_ORIGIN = os.getenv("CLIENT_ORIGIN", "unknown_origin")


EVENT_DOCUMENT_DIR = "uploads/event/documents"
if not os.path.exists(EVENT_DOCUMENT_DIR):
    os.makedirs(EVENT_DOCUMENT_DIR)

EVENT_BANNER_DIR = "uploads/event/banners"
if not os.path.exists(EVENT_BANNER_DIR):
    os.makedirs(EVENT_BANNER_DIR)

ORG_UNIT_LOGO_DIR = "uploads/org_unit/logos"
if not os.path.exists(ORG_UNIT_LOGO_DIR):
    os.makedirs(ORG_UNIT_LOGO_DIR)


def get_object(id: int, db: Session, model):
    data = db.query(model).filter(model.id == id).first()
    if data is None:
        raise HTTPException(
            status_code=404,
            detail=f"{model.__name__} with ID {id} does not exist or has been deleted",
        )
    return data


def sanitize_filename(name: str) -> str:
    # Normalize to remove accents
    name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    # Replace spaces and remove non-word characters
    name = re.sub(r"[^\w\s-]", "", name).strip().replace(" ", "_")
    return name


# Define mapping of participation_role keys to display names
PARTICIPATION_ROLE_MAP = {
    "secretariat": "ECSA-HC secretariat",
    "djcc": "DJCC Member",
    "moh": "Country delegate (from Ministry of Health)",
    "member_state": "Participant from ECSA Member States",
    "other_africa": "Participant from other African countries",
    "world": "Participant from the Rest of the World",
    "student": "Student",
    "exibitor": "Sponsor/Exhibitor",
}


def convert_png_to_rgb(path):
    img = Image.open(path)
    if img.mode in ("RGBA", "LA"):
        background = Image.new("RGB", img.size, (255, 255, 255))
        background.paste(img, mask=img.split()[3])  # Use alpha channel as mask
        return ImageReader(background)
    return ImageReader(img)


def load_logo_with_transparency(path):
    """Load a PNG preserving its alpha channel, for drawing with mask='auto'
    so the badge's textured background shows through instead of a white box."""
    img = Image.open(path)
    if img.mode != "RGBA":
        img = img.convert("RGBA")
    return ImageReader(img)


def normalize_event_name(name: str) -> str:
    return (
        name.replace("ᵗʰ", "th")
        .replace("ˢᵗ", "st")
        .replace("ⁿᵈ", "nd")
        .replace("ʳᵈ", "rd")
    )


# ── Badge role colours / labels (matching official ECSA name-tag templates) ───
BADGE_ROLE_COLORS = {
    "media":        "#FFD700",
    "moderator":    "#F7941D",
    "secretariat":  "#00AEEF",
    "speaker":      "#C8102E",
    "presenter":    "#C8102E",
    "delegate":     "#009639",
    "djcc":         "#009639",
    "moh":          "#009639",
    "member_state": "#009639",
    "other_africa": "#009639",
    "world":        "#009639",
    "student":      "#009639",
    "participant":  "#009639",
    "exhibitor":    "#F7941D",
    "sponsor":      "#F7941D",
    "local_secretariat": "#00AEEF",
    "usher":             "#EC4899",
    "driver":            "#92400E",
    "medical_staff":     "#0D9488",
}

BADGE_ROLE_LABELS = {
    "media":        "MEDIA",
    "moderator":    "MODERATOR",
    "secretariat":  "SECRETARIAT",
    "speaker":      "SPEAKER",
    "presenter":    "PRESENTER",
    "delegate":     "DELEGATE",
    "djcc":         "DELEGATE",
    "moh":          "DELEGATE",
    "member_state": "DELEGATE",
    "other_africa": "DELEGATE",
    "world":        "DELEGATE",
    "student":      "STUDENT",
    "participant":  "PARTICIPANT",
    "exhibitor":    "EXHIBITOR",
    "sponsor":      "SPONSOR",
    "local_secretariat": "SECRETARIAT",
    "usher":             "USHER",
    "driver":            "DRIVER",
    "medical_staff":     "MEDICAL STAFF",
}

# Roles exempt from payment (auto paid=True at creation, no payment reminders).
# Support-staff categories (ushers/drivers/medical staff/local secretariat) are
# never expected to pay like delegates — matches how "secretariat" already works.
NO_PAYMENT_ROLES = {"secretariat", "local_secretariat", "usher", "driver", "medical_staff"}

# "local_secretariat" is an umbrella role_category filter covering the whole
# local support team — ushers/drivers/medical staff included. Each still gets
# its own distinct badge color/label; this only groups them for filtering.
LOCAL_SECRETARIAT_ROLES = {"local_secretariat", "usher", "driver", "medical_staff"}

# Generic word abbreviations used to shorten organization names that don't fit
# their badge box even at minimum font size (must match ParticipantBadgeModal.vue).
# Long country/region names go first since they're multi-word phrases.
ORG_ABBREVIATIONS = [
    (r"\bUnited Kingdom\b", "UK"), (r"\bUnited States of America\b", "USA"),
    (r"\bUnited States\b", "US"), (r"\bUnited Arab Emirates\b", "UAE"),
    (r"\bDemocratic Republic of (the )?Congo\b", "DRC"),
    (r"\bUniversity\b", "Univ."), (r"\bInstitute\b", "Inst."),
    (r"\bDepartment\b", "Dept."), (r"\bMinistry\b", "Min."),
    (r"\bInternational\b", "Int'l"), (r"\bOrgani[sz]ation\b", "Org."),
    (r"\bAssociation\b", "Assoc."), (r"\bFoundation\b", "Fdn."),
    (r"\bCorporation\b", "Corp."), (r"\bCompany\b", "Co."),
    (r"\bLimited\b", "Ltd."), (r"\bGovernment\b", "Govt."),
    (r"\bNational\b", "Natl."), (r"\bRegional\b", "Reg'l"),
    (r"\bProgramme\b", "Prog."), (r"\bProgram\b", "Prog."),
    (r"\bManagement\b", "Mgmt."), (r"\bDevelopment\b", "Dev."),
    (r"\bCommunity\b", "Cmty."), (r"\bRepublic\b", "Rep."),
    (r"\bAfrican\b", "Afr."), (r"\bSouthern\b", "S."),
    (r"\bEastern\b", "E."), (r"\bWestern\b", "W."),
    (r"\bNorthern\b", "N."), (r"\bCentral\b", "Ctrl."),
    (r"\bHealth\b", "Hlth"), (r"\bServices\b", "Svcs"),
    (r"\bAgency\b", "Agcy"), (r"\bAuthority\b", "Auth."),
    (r"\bCommission\b", "Comm."), (r"\bResearch\b", "Rsch"),
    (r"\bTechnology\b", "Tech."), (r"\bAdministration\b", "Admin."),
]

# Words skipped when generating an initials-based acronym fallback (e.g.
# "University College London" -> "UCL") for orgs with no "(ACRONYM)" of
# their own (must match ORG_STOPWORDS in ParticipantBadgeModal.vue).
ORG_STOPWORDS = {"and", "of", "for", "the", "in", "on", "at", "a", "an", "&", "to"}


def abbreviate_org_words(text):
    out = text
    for pattern, repl in ORG_ABBREVIATIONS:
        out = re.sub(pattern, repl, out, flags=re.IGNORECASE)
    return out


def generate_org_acronym(text):
    """Initials of the significant words before the last comma (kept as a
    trailing country/location suffix), e.g. "National Health Insurance
    Management Authority, Zambia" -> "NHIMA, Zambia". Returns None if too
    short to be a meaningful acronym."""
    main, _, suffix = text.partition(",")
    main = re.sub(r"\([^)]*\)", "", main)
    words = re.findall(r"[A-Za-z']+", main)
    initials = "".join(w[0].upper() for w in words if w.lower() not in ORG_STOPWORDS)
    if len(initials) < 2:
        return None
    suffix = suffix.strip()
    return f"{initials}, {suffix}" if suffix else initials


# ISO 3166-1 alpha-2 codes for ECSA member states shown on the badge
# (must match flagCodes in ParticipantBadgeModal.vue)
ECSA_FLAG_CODES = ["sz", "ke", "ls", "mw", "mu", "mz", "st", "tz", "ug", "zm", "zw"]
_FLAG_CACHE: dict = {}


def _get_flag_images():
    """Return a {code: ImageReader} dict for ECSA member flags (cached)."""
    global _FLAG_CACHE
    if _FLAG_CACHE:
        return _FLAG_CACHE
    flags_dir = "assets/flags"
    os.makedirs(flags_dir, exist_ok=True)
    from urllib.request import urlretrieve
    result = {}
    for code in ECSA_FLAG_CODES:
        local = f"{flags_dir}/{code}.png"
        if not os.path.exists(local):
            try:
                urlretrieve(f"https://flagcdn.com/40x30/{code}.png", local)
            except Exception:
                result[code] = None
                continue
        try:
            result[code] = convert_png_to_rgb(local)
        except Exception:
            result[code] = None
    _FLAG_CACHE = result
    return result


def _fmt_event_dates(event) -> str:
    """Return a human-readable date range for an event."""
    def _fmt(d):
        try:
            return d.strftime("%-d %B %Y")
        except Exception:
            return str(d)
    if event.start_date and event.end_date:
        return f"{_fmt(event.start_date)} – {_fmt(event.end_date)}"
    if event.start_date:
        return _fmt(event.start_date)
    return ""


@router.get("")
@router.get("/")
async def get_events(
    request: Request,
    db: Session = Depends(get_db),
    skip: int = Query(default=0, ge=0),
    limit: int = 10,
    search: str = "",
    dependency: Dependency = Depends(get_dependency),
):
    client_ip = dependency.request_ip(request)
    dependency.log_activity(None, "VIEW_EVENTS", "None", client_ip, "Get all events")

    search_filter = or_(
        Event.event.ilike(f"%{search}%"),
        Event.theme.ilike(f"%{search}%"),
        Event.description.ilike(f"%{search}%"),
    )

    filters = [Event.deleted_at.is_(None)]
    if search_filter is not None:
        filters.insert(0, search_filter)

    events_query = db.query(Event).options(joinedload(Event.org_unit)).filter(*filters)

    total_count = events_query.count()
    events = events_query.offset(skip).limit(limit).all()

    pages = math.ceil(total_count / limit)
    return {
        "pages": pages,
        "data": [
            {
                "id": e.id,
                "event": e.event,
                "theme": e.theme,
                "description": e.description,
                "start_date": e.start_date,
                "end_date": e.end_date,
                "location": e.location,
                "banner_image": e.banner_image,
                "organizers": e.organizers,
                "org_unit_id": e.org_unit_id,
                "org_unit": {
                    "id": e.org_unit.id,
                    "name": e.org_unit.name,
                    "primary_color": e.org_unit.primary_color or "#0095B6",
                    "secondary_color": e.org_unit.secondary_color or "#F7941D",
                    "logo": e.org_unit.logo,
                } if e.org_unit else None,
                "country_id": e.country_id,
            }
            for e in events
        ],
    }


@router.post("/")
async def add_event(
    request: Request,
    event_schema: EventSchema,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    auth_dependency.secure_access("ADD_EVENT", current_user["user_id"])

    client_ip = dependency.request_ip(request)
    dependency.log_activity(
        current_user["user_id"],
        "ADD_EVENT",
        current_user["username"],
        client_ip,
        event_schema.event,
    )

    create_event_model = Event(
        org_unit_id=event_schema.org_unit_id,
        country_id=event_schema.country_id,
        user_id=current_user["user_id"],
        event=event_schema.event,
        theme=event_schema.theme,
        description=event_schema.description,
        location=event_schema.location,
        start_date=event_schema.start_date,
        end_date=event_schema.end_date,
        banner_image=event_schema.banner_image,
        organizers=event_schema.organizers,
        participation_info=event_schema.participation_info,
        logistics_info=event_schema.logistics_info,
        sponsors_info=event_schema.sponsors_info,
    )

    db.add(create_event_model)
    db.commit()
    db.refresh(create_event_model)
    return {"id": create_event_model.id}


def _build_pending_list(pending_regs, db, event_id=None):
    """Build pending registration list flagging accepted abstract authors and whether they were reminded."""
    from sqlalchemy import text as _t
    from models.models import EmailLog

    # Accepted abstract author emails for this event
    abstract_author_emails = set()
    if event_id:
        rows = db.execute(_t(
            "SELECT DISTINCT LOWER(aa.email) FROM abstract a "
            "JOIN abstract_author aa ON aa.abstract_id = a.id "
            "WHERE a.event_id = :eid AND a.status = 'accepted' AND aa.email IS NOT NULL"
        ), {"eid": event_id}).fetchall()
        abstract_author_emails = {row[0] for row in rows}

    # Emails that have already received a registration_reminder
    reminded_emails = {
        row.recipient_email.lower()
        for row in db.query(EmailLog.recipient_email)
        .filter(EmailLog.email_type == "registration_reminder", EmailLog.status == "sent")
        .all()
    }

    return [
        {
            "id": r.id,
            "user_id": r.user_id,
            "firstname": r.user.firstname if r.user else None,
            "lastname": r.user.lastname if r.user else None,
            "email": r.user.email if r.user else None,
            "phone": r.user.phone if r.user else None,
            "country": (
                r.badge_country
                or (r.user.user_profile[0].country.country
                    if r.user and r.user.user_profile and r.user.user_profile[0].country
                    else None)
            ),
            "participation_role": r.participation_role,
            "registered_at": r.registered_at,
            "notes": getattr(r, "notes", None),
            "is_abstract_author": (r.user.email or "").lower() in abstract_author_emails,
            "abstract_reminder_sent": (r.user.email or "").lower() in reminded_emails,
        }
        for r in pending_regs
    ]


# ══════════════════════════════════════════════════════════════════════════════
#  EVENT TEMPLATES  (PPT / poster templates downloadable by accepted+paid users)
#  NOTE: must stay before the generic /{event_id} routes below, otherwise
#  FastAPI matches "/templates" etc. against /{event_id} first (registration
#  order, not specificity) and 422s trying to parse "templates" as an int.
# ══════════════════════════════════════════════════════════════════════════════

TEMPLATE_DIR = "uploads/templates"
if not os.path.exists(TEMPLATE_DIR):
    os.makedirs(TEMPLATE_DIR)


@router.post("/templates/upload", status_code=201)
async def upload_event_template(
    file: UploadFile = File(...),
    event_id: int = Form(None),
    name: str = Form(...),
    presentation_type: str = Form(None),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(lambda db=Depends(get_db): Auth(db)),
):
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])
    from models.models import EventTemplate
    ext = os.path.splitext(file.filename)[1]
    unique_name = f"{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(TEMPLATE_DIR, unique_name)
    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    t = EventTemplate(
        event_id=event_id or None,
        name=name,
        file_path=file_path,
        presentation_type=presentation_type or None,
        uploaded_by=current_user["user_id"],
    )
    db.add(t)
    db.commit()
    db.refresh(t)
    return _serialize_template(t)


@router.get("/templates")
def list_templates(
    event_id: int = Query(None),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(lambda db=Depends(get_db): Auth(db)),
):
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])
    from models.models import EventTemplate
    q = db.query(EventTemplate)
    if event_id:
        q = q.filter(
            (EventTemplate.event_id == event_id) | (EventTemplate.event_id == None)
        )
    return [_serialize_template(t) for t in q.order_by(EventTemplate.uploaded_at.desc()).all()]


class TemplateUpdateSchema(BaseModel):
    name: Optional[str] = None
    event_id: Optional[int] = None
    presentation_type: Optional[str] = None


@router.patch("/templates/{template_id}")
def update_template(
    template_id: int,
    schema: TemplateUpdateSchema,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(lambda db=Depends(get_db): Auth(db)),
):
    """Correct a template's name/event/type without deleting and re-uploading the file."""
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])
    from models.models import EventTemplate
    t = db.query(EventTemplate).filter(EventTemplate.id == template_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    if schema.name is not None:
        t.name = schema.name
    # event_id/presentation_type: 0 or "" from the frontend means "clear to All"
    if schema.event_id is not None:
        t.event_id = schema.event_id or None
    if schema.presentation_type is not None:
        t.presentation_type = schema.presentation_type or None
    db.commit()
    db.refresh(t)
    return _serialize_template(t)


@router.delete("/templates/{template_id}", status_code=204)
def delete_template(
    template_id: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(lambda db=Depends(get_db): Auth(db)),
):
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])
    from models.models import EventTemplate
    t = db.query(EventTemplate).filter(EventTemplate.id == template_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    if os.path.exists(t.file_path):
        os.remove(t.file_path)
    db.delete(t)
    db.commit()


def _template_eligible_recipients(template, db: Session, event_id: int = None):
    """Accepted-abstract authors, paid for the given event, whose abstract's
    presentation_type matches the template (or any type if the template has none
    set). Templates scoped to a specific event use that event; "all events"
    templates require the caller to pass event_id explicitly — dispatch always
    targets one event so we never mass-notify across every event at once."""
    target_event_id = template.event_id or event_id
    if not target_event_id:
        return {}
    from models.models import Abstract, AbstractStatus, PresentationType

    abstracts_q = db.query(Abstract).options(joinedload(Abstract.authors)).filter(
        Abstract.event_id == target_event_id,
        Abstract.status == AbstractStatus.accepted,
        Abstract.deleted_at == None,
    )
    if template.presentation_type:
        abstracts_q = abstracts_q.filter(Abstract.presentation_type == PresentationType(template.presentation_type))
    abstracts = abstracts_q.all()

    # Eligible on payment: fully verified paid, or at least proof of payment
    # submitted (admin just hasn't verified it yet) — don't lock someone out of
    # their own accepted-abstract presentation just because verification is pending.
    registrations = db.query(Registration).filter(
        Registration.event_id == target_event_id,
        or_(Registration.paid == True, Registration.payment_proof.isnot(None)),
    ).all()
    payment_status_by_user_id = {
        r.user_id: ("paid" if r.paid else "proof_uploaded") for r in registrations
    }
    payment_status_by_email = {
        u.email.lower(): payment_status_by_user_id[u.id]
        for u in db.query(User).filter(User.id.in_(payment_status_by_user_id.keys())).all()
        if u.email
    } if payment_status_by_user_id else {}

    by_email = {}
    for a in abstracts:
        firstname, email = None, None
        if a.authors:
            presenting = [au for au in a.authors if au.is_presenting and au.email]
            candidates = presenting or [au for au in a.authors if au.email][:1]
            if candidates:
                firstname, email = candidates[0].firstname or "Presenter", candidates[0].email
        if not email and a.submitted_by:
            submitter = db.query(User).filter(User.id == a.submitted_by).first()
            if submitter and submitter.email:
                firstname, email = submitter.firstname or "Presenter", submitter.email
        payment_status = payment_status_by_email.get((email or "").lower())
        if not email or not payment_status:
            continue
        key = email.lower()
        by_email.setdefault(key, {"firstname": firstname or "Presenter", "email": email, "payment_status": payment_status, "abstracts": []})
        by_email[key]["abstracts"].append(a)
    return by_email


@router.get("/templates/{template_id}/notify-preview")
def template_notify_preview(
    template_id: int,
    event_id: int = Query(None),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(lambda db=Depends(get_db): Auth(db)),
):
    """Preview who would receive a 'template available' notification (no emails sent)."""
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])
    from models.models import EventTemplate, EmailLog

    template = db.query(EventTemplate).filter(EventTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    if not template.event_id and not event_id:
        return {
            "template_name": template.name, "event_name": None,
            "to_send": [], "already_notified": [], "total_recipients": 0,
            "needs_event": True,
            "events": [{"id": e.id, "event": e.event} for e in db.query(Event).order_by(Event.start_date.desc()).all()],
        }

    target_event_id = template.event_id or event_id
    event = db.query(Event).filter(Event.id == target_event_id).first()
    by_email = _template_eligible_recipients(template, db, event_id=target_event_id)

    notified_emails = {
        row.recipient_email.lower()
        for row in db.query(EmailLog.recipient_email).filter(
            EmailLog.email_type == f"presentation_template_{template_id}_{target_event_id}",
            EmailLog.status == "sent",
        ).all()
    }

    to_send, already_notified = [], []
    for entry in by_email.values():
        row = {
            "firstname": entry["firstname"], "email": entry["email"],
            "abstract_titles": [a.title for a in entry["abstracts"]],
            # Every recipient here already passed the accepted + (paid or proof-uploaded)
            # filter above — surfaced explicitly so the admin can verify it, not just trust it.
            "status": "accepted",
            "payment_status": entry["payment_status"],
        }
        (already_notified if entry["email"].lower() in notified_emails else to_send).append(row)

    return {
        "template_name": template.name,
        "event_name": event.event if event else None,
        "to_send": to_send,
        "already_notified": already_notified,
        "total_recipients": len(by_email),
    }


class NotifyTemplateSchema(BaseModel):
    event_id: Optional[int] = None
    test_email: Optional[str] = None


@router.post("/templates/{template_id}/notify")
def template_notify(
    template_id: int,
    schema: NotifyTemplateSchema,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(lambda db=Depends(get_db): Auth(db)),
):
    """Email paid, accepted presenters (matching the template's type) that a
    presentation template is now available in their account."""
    auth_dependency.secure_access("VIEW_ABSTRACTS", current_user["user_id"])
    from models.models import EventTemplate, EmailLog
    import utils.mailer_util as _mailer

    template = db.query(EventTemplate).filter(EventTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    if not template.event_id and not schema.event_id:
        raise HTTPException(status_code=400, detail="Choose an event to notify presenters for this template")

    target_event_id = template.event_id or schema.event_id
    event = db.query(Event).filter(Event.id == target_event_id).first()
    by_email = _template_eligible_recipients(template, db, event_id=target_event_id)

    email_type = f"presentation_template_{template_id}_{target_event_id}"
    notified_emails = {
        row.recipient_email.lower()
        for row in db.query(EmailLog.recipient_email).filter(
            EmailLog.email_type == email_type, EmailLog.status == "sent",
        ).all()
    }
    jobs = [v for v in by_email.values() if v["email"].lower() not in notified_emails]

    if schema.test_email:
        # A dry run always sends exactly one email, regardless of recipient count.
        jobs = jobs[:1] or [{"firstname": "Test", "email": schema.test_email, "abstracts": []}]
    elif not jobs:
        return {"sent": 0, "message": "No unnotified eligible presenters found for this template."}

    subject = f"Presentation Template Available – {event.event if event else ''}"
    ptype = template.presentation_type or "either"

    messages = []
    for j in jobs:
        recipient = schema.test_email or j["email"]
        body = _mailer.templates.get_template("presentation_template_notify.html").render(
            subject=subject,
            firstname=j["firstname"],
            event_name=event.event if event else "",
            abstract_titles=[a.title for a in j["abstracts"]],
            presentation_type=ptype,
            portal_url=_mailer.APP_BASE_URL,
            year=_mailer.YEAR,
        )
        messages.append({
            "recipient_email": recipient,
            "subject": subject,
            "body": body,
            "email_type": email_type if not schema.test_email else "presentation_template_test",
            "sent_by_user_id": current_user["user_id"],
        })

    background_tasks.add_task(_mailer.send_bulk_emails, messages, db)

    return {
        "sent": len(messages),
        "message": (
            f"Notification queued for {len(messages)} presenter(s)."
            if not schema.test_email else "Test email queued."
        ),
    }


@router.get("/templates/for-me")
def my_templates(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Return templates accessible to this user (accepted abstract + paid registration)."""
    from models.models import EventTemplate, Abstract, AbstractStatus, Registration
    user_id = current_user["user_id"]

    # Find event IDs where user has a paid registration
    paid_event_ids = {
        r.event_id for r in
        db.query(Registration).filter(
            Registration.user_id == user_id, Registration.paid == True
        ).all()
    }

    # Find event IDs where user has an accepted abstract (submitted_by OR as author)
    from models.models import AbstractAuthor
    from sqlalchemy import or_
    accepted_abstracts = db.query(Abstract).filter(
        Abstract.status == AbstractStatus.accepted,
        Abstract.deleted_at == None,
        or_(
            Abstract.submitted_by == user_id,
            Abstract.id.in_(
                db.query(AbstractAuthor.abstract_id)
                .join(User, User.email == AbstractAuthor.email)
                .filter(User.id == user_id)
            )
        )
    ).all()
    accepted_event_ids = {a.event_id for a in accepted_abstracts}

    # Eligible = paid AND accepted
    eligible_event_ids = paid_event_ids & accepted_event_ids

    if not eligible_event_ids:
        return []

    from models.models import EventTemplate
    templates = db.query(EventTemplate).filter(
        (EventTemplate.event_id.in_(eligible_event_ids)) |
        (EventTemplate.event_id == None)
    ).order_by(EventTemplate.uploaded_at.desc()).all()

    # Also attach presentation_type from the user's abstract
    abs_type_by_event = {}
    for a in accepted_abstracts:
        abs_type_by_event[a.event_id] = a.presentation_type.value if a.presentation_type else None

    result = []
    for t in templates:
        p_type = abs_type_by_event.get(t.event_id) if t.event_id else None
        # Only show template if ptype matches user's abstract type (or template has no type restriction)
        if t.presentation_type and p_type and t.presentation_type != p_type:
            continue
        result.append(_serialize_template(t))
    return result


def _serialize_template(t):
    from models.models import Event as _Event
    return {
        "id": t.id,
        "event_id": t.event_id,
        "name": t.name,
        "file_path": t.file_path,
        "url": f"/uploads/templates/{os.path.basename(t.file_path)}",
        "presentation_type": t.presentation_type,
        "uploaded_at": t.uploaded_at,
    }


@router.get("/{event_id}/basic")
async def get_event_basic(
    event_id: int,
    db: Session = Depends(get_db),
):
    """Lightweight event lookup for public pages (e.g. QR attendance check-in)
    that only need the name, dates, location, and brand colors — not the full
    participants/pending-registrations/abstract-stats payload from GET /{event_id},
    which is expensive on events with hundreds of registrations."""
    event = get_object(event_id, db, Event)
    if not event:
        raise HTTPException(status_code=404, detail="event not found")
    return {
        "event": {
            "id": event.id,
            "event": event.event,
            "location": event.location,
            "start_date": event.start_date,
            "end_date": event.end_date,
            "org_unit_primary_color": event.org_unit.primary_color if event.org_unit else "#0095B6",
            "org_unit_secondary_color": event.org_unit.secondary_color if event.org_unit else "#F7941D",
        }
    }


@router.get("/{event_id}")
async def get_event(
    request: Request,
    event_id: int,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    current_user: Optional[dict] = Depends(get_optional_current_user),
):
    client_ip = dependency.request_ip(request)

    dependency.log_activity(
        current_user["user_id"] if current_user else None,
        "VIEW_EVENT",
        "None",
        client_ip,
        f"View event id {event_id} and associated permissions",
    )

    # Eager-load registrations + their user/profile/photo in bulk instead of
    # lazily (one query per participant, per relationship hop) — with 300+
    # registrations that was ~1000 extra round trips and made this endpoint
    # crawl. selectinload issues one extra query per relationship level
    # regardless of how many rows it covers.
    event = (
        db.query(Event)
        .options(
            joinedload(Event.country),
            joinedload(Event.org_unit),
            joinedload(Event.user),
            selectinload(Event.documents),
            selectinload(Event.links),
            selectinload(Event.registrations)
                .selectinload(Registration.user)
                .selectinload(User.user_profile)
                .joinedload(UserProfile.country),
            selectinload(Event.registrations)
                .selectinload(Registration.user)
                .selectinload(User.user_photo),
        )
        .filter(Event.id == event_id, Event.deleted_at == None)
        .first()
    )
    if not event:
        raise HTTPException(
            status_code=404,
            detail=f"Event with ID {event_id} does not exist or has been deleted",
        )
    if event:
        _all_regs = [r for r in (event.registrations or []) if r.deleted_at is None]
        # Confirmed = paid OR has uploaded proof; pending = neither
        registrations = [r for r in _all_regs if r.paid or r.payment_proof]
        pending_payment_regs = [r for r in _all_regs if not r.paid and not r.payment_proof]
        documents = event.documents or []
        links = event.links or []

        # ── Determine the current user's access level ─────────────────────────
        # "none"   → not logged in
        # "unpaid" → logged in but no paid registration for this event
        # "paid"   → has a paid registration OR has admin/secretariat permission
        user_access = "none"
        is_admin = False
        if current_user:
            uid = current_user["user_id"]
            # Check for admin permission (ADMIN_DASHBOARD) → always full access
            from models.models import UserRole, RolePermission, Permission as Perm
            is_admin = db.query(UserRole).join(
                RolePermission, UserRole.role_id == RolePermission.role_id
            ).join(
                Perm, RolePermission.permission_id == Perm.id
            ).filter(
                UserRole.user_id == uid,
                Perm.permission_code == "ADMIN_DASHBOARD",
            ).first()

            if is_admin:
                user_access = "paid"
            else:
                reg = db.query(Registration).filter(
                    Registration.user_id == uid,
                    Registration.event_id == event_id,
                    Registration.deleted_at == None,
                ).first()
                if reg and reg.paid:
                    user_access = "paid"
                else:
                    user_access = "unpaid"

        # Fetch payment data via raw SQL to avoid ORM enum conversion errors on legacy data
        from sqlalchemy import text as _sql_text

        # Abstract author registration stats
        _abstract_author_stats = {"total_authors": 0, "registered": 0, "not_registered": 0}
        try:
            _author_rows = db.execute(
                _sql_text(
                    "SELECT DISTINCT LOWER(aa.email) as email"
                    " FROM abstract a"
                    " JOIN abstract_author aa ON aa.abstract_id = a.id"
                    " WHERE a.event_id = :eid AND a.status = 'accepted' AND aa.email IS NOT NULL"
                ),
                {"eid": event_id},
            ).fetchall()
            _total_authors = len(_author_rows)
            if _total_authors:
                _registered_count = db.execute(
                    _sql_text(
                        "SELECT COUNT(DISTINCT LOWER(u.email)) as cnt"
                        " FROM registration r"
                        " JOIN user u ON u.id = r.user_id"
                        " WHERE r.event_id = :eid AND r.deleted_at IS NULL"
                        " AND LOWER(u.email) IN ("
                        "   SELECT DISTINCT LOWER(aa2.email)"
                        "   FROM abstract a2"
                        "   JOIN abstract_author aa2 ON aa2.abstract_id = a2.id"
                        "   WHERE a2.event_id = :eid AND a2.status = 'accepted' AND aa2.email IS NOT NULL"
                        " )"
                    ),
                    {"eid": event_id},
                ).scalar() or 0
                _abstract_author_stats = {
                    "total_authors": _total_authors,
                    "registered": int(_registered_count),
                    "not_registered": _total_authors - int(_registered_count),
                }
        except Exception as _ae:
            logger.error(f"Abstract author stats error: {_ae}")
        _reg_ids = [r.id for r in registrations]
        if _reg_ids:
            _placeholders = ",".join(str(rid) for rid in _reg_ids)
            _payment_rows = db.execute(
                _sql_text(
                    f"SELECT registration_id, payment_method, payment_amount, payment_date"
                    f" FROM payment WHERE registration_id IN ({_placeholders})"
                )
            ).fetchall()
            payment_by_reg = {row.registration_id: row for row in _payment_rows}
        else:
            payment_by_reg = {}

        # This endpoint is hit by both the admin dashboard (needs everyone)
        # and ordinary logged-in users viewing their own event page (needs
        # only their own record) — auth is optional, so without this gate
        # anyone, logged in or not, could pull every paid participant's
        # phone, email, exact payment amount/date, payment proof file path,
        # and admin notes just by requesting this endpoint directly.
        _all_participants_serialized = [
            {
                "id": r.id,
                "user_id": r.user_id,
                "firstname": r.user.firstname if r.user else None,
                "lastname": r.user.lastname if r.user else None,
                "phone": r.user.phone if r.user else None,
                "email": r.user.email if r.user else None,
                "photo": (
                    r.user.user_photo if r.user and r.user.user_photo else None
                ),
                "country_id": (
                    r.user.user_profile[0].country_id
                    if r.user and r.user.user_profile
                    else None
                ),
                # badge_* fields (set by admin bulk-import/add) take priority over a
                # general profile when both exist — e.g. someone re-imported for a
                # different event with fresh title/org data shouldn't have it
                # silently overridden by an older, unrelated profile.
                "country": (
                    r.badge_country
                    or (r.user.user_profile[0].country.country
                        if r.user and r.user.user_profile and r.user.user_profile[0].country
                        else None)
                ),
                "participation_role": r.participation_role,
                "organisation": (
                    r.badge_organisation
                    or (r.user.user_profile[0].organisation
                        if r.user and r.user.user_profile else None)
                ),
                "position": (
                    r.badge_position
                    or (r.user.user_profile[0].position
                        if r.user and r.user.user_profile else None)
                ),
                "title": (
                    r.badge_prefix
                    or (r.user.user_profile[0].title
                        if r.user and r.user.user_profile else None)
                ),
                "paid": getattr(r, "paid", None),
                "payment_proof": getattr(r, "payment_proof", None),
                "notes": getattr(r, "notes", None),
                "payment_method": payment_by_reg[r.id].payment_method if r.id in payment_by_reg else None,
                "payment_amount": float(payment_by_reg[r.id].payment_amount) if r.id in payment_by_reg and payment_by_reg[r.id].payment_amount else None,
                "payment_date": str(payment_by_reg[r.id].payment_date) if r.id in payment_by_reg and payment_by_reg[r.id].payment_date else None,
                "registered_at": r.registered_at,
                "updated_at": r.updated_at,
                "reminder_sent_at": getattr(r, "reminder_sent_at", None),
                "badge_exported_at": getattr(r, "badge_exported_at", None),
            }
            for r in registrations
        ]
        if is_admin:
            _visible_participants = _all_participants_serialized
            _visible_pending = _build_pending_list(pending_payment_regs, db, event_id=event_id)
        elif current_user:
            _visible_participants = [
                p for p in _all_participants_serialized if p["user_id"] == current_user["user_id"]
            ]
            _visible_pending = []
        else:
            _visible_participants = []
            _visible_pending = []

        return {
            "event": {
                "id": event.id,
                "event": event.event,
                "country_id": event.country_id,
                "country": event.country.country if event.country else None,
                "org_unit_id": event.org_unit_id,
                "org_unit": event.org_unit.name if event.org_unit else None,
                "org_unit_primary_color": event.org_unit.primary_color if event.org_unit else "#0095B6",
                "org_unit_secondary_color": event.org_unit.secondary_color if event.org_unit else "#F7941D",
                "org_unit_logo": event.org_unit.logo if event.org_unit else None,
                "user_id": event.user_id,
                "firstname": event.user.firstname if event.user else None,
                "lastname": event.user.lastname if event.user else None,
                "location": event.location,
                "theme": event.theme,
                "description": event.description,
                "start_date": event.start_date,
                "end_date": event.end_date,
                "banner_image": event.banner_image,
                "organizers": event.organizers,
                "participation_info": event.participation_info,
                "logistics_info": event.logistics_info,
                "sponsors_info": event.sponsors_info,
                "participation_role": (
                    registrations[0].participation_role if registrations else None
                ),
                "user_access": user_access,
            },
            "participants": _visible_participants,
            "documents": [
                {
                    "id": d.id,
                    "document_type": d.document_type,
                    "file_type": d.file_type,
                    "file_name": d.file_name,
                    "name": d.name,
                    "path": d.path,
                    "access_level": d.access_level,
                }
                for d in documents
            ],
            "links": [
                {
                    "id": l.id,
                    "name": l.name,
                    "link": l.link,
                }
                for l in links
            ],
            "pending_registrations": _visible_pending,
            "abstract_author_stats": _abstract_author_stats,
        }


@router.put("/{event_id}")
async def update_event(
    request: Request,
    event_id: int,
    current_user: user_dependency,
    event_schema: EventUpdateSchema,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    auth_dependency.secure_access("UPDATE_EVENT", current_user["user_id"])

    client_ip = dependency.request_ip(request)

    dependency.log_activity(
        current_user["user_id"],
        "UPDATE_EVENT",
        current_user["username"],
        client_ip,
        f"Update event id {event_id}",
    )
    event_model = get_object(event_id, db, Event)

    event_model.org_unit_id = event_schema.org_unit_id
    event_model.country_id = event_schema.country_id
    event_model.user_id = current_user["user_id"]
    event_model.event = event_schema.event
    event_model.theme = event_schema.theme
    event_model.description = event_schema.description
    event_model.location = event_schema.location
    event_model.start_date = event_schema.start_date
    event_model.end_date = event_schema.end_date
    event_model.organizers = event_schema.organizers
    event_model.participation_info = event_schema.participation_info
    event_model.logistics_info = event_schema.logistics_info
    event_model.sponsors_info = event_schema.sponsors_info
    if event_schema.banner_image is not None:
        event_model.banner_image = event_schema.banner_image

    db.commit()
    db.refresh(event_model)
    return event_schema


@router.delete("/{event_id}")
async def delete_event(
    request: Request,
    event_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    auth_dependency.secure_access("DELETE_EVENT", current_user["user_id"])

    dependency.cascade_soft_delete_recursive(Event, event_id)

    client_ip = dependency.request_ip(request)
    event = get_object(event_id, db, Event)

    client_ip = dependency.request_ip(request)
    dependency.log_activity(
        current_user["user_id"],
        "DELETE_EVENT",
        current_user["username"],
        client_ip,
        f"Delete event id {event_id} event {event.event}",
    )
    return {"detail": "event Successfully deleted"}


@router.get("/registration/{registration_id}")
async def get_registration_details(
    request: Request,
    registration_id: int,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
):
    client_ip = dependency.request_ip(request)

    dependency.log_activity(
        1,
        "VIEW_REGISTRATION",
        "None",
        client_ip,
        f"View registration ID {registration_id}",
    )

    registration = get_object(registration_id, db, Registration)
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    user = db.query(User).filter(User.id == registration.user_id).first()
    event = db.query(Event).filter(Event.id == registration.event_id).first()

    role_value = (
        registration.participation_role.name
        if hasattr(registration.participation_role, "name")
        else str(registration.participation_role).lower()
    )

    return {
        "registration": {
            "id": registration.id,
            "user_id": registration.user_id,
            "event_id": registration.event_id,
            "participation_role": role_value,
            "paid": registration.paid,
            "created_at": registration.created_at,
            "updated_at": registration.updated_at,
        },
        "user": (
            {
                "id": user.id,
                "firstname": user.firstname,
                "lastname": user.lastname,
                "email": user.email,
                "phone": user.phone,
            }
            if user
            else None
        ),
        "event": (
            {
                "id": event.id,
                "event": event.event,
                "location": event.location,
                "start_date": event.start_date,
                "end_date": event.end_date,
            }
            if event
            else None
        ),
    }


# Maps frontend strings → PaymentMethod enum name (accessed via PaymentMethod[name])
METHOD_MAP = {
    "bank transfer": "BANK_TRANSFER",
    "mpesa": "MPESA",
    "cash": "CASH",
    "card": "CARD",
    "mobile money": "MPESA",
    "credit card": "CARD",
    "debit card": "CARD",
    "online payment (credit/debit card)": "CARD",
    "online": "CARD",
}


@router.post("/payment/")
async def submit_payment(
    registration_id: int = Form(...),
    event_id: int = Form(...),
    payment_method: str = Form(...),
    payment_amount: float = Form(...),
    proof_file: Optional[UploadFile] = File(None),
    background_tasks: BackgroundTasks = None,
    db: Session = Depends(get_db),
):
    from models.models import PaymentMethod, PaymentStatus
    from utils.mailer_util import proof_of_payment_received_email

    registration = get_object(registration_id, db, Registration)

    raw_method = payment_method.strip()
    enum_name = METHOD_MAP.get(raw_method.lower())
    if not enum_name:
        raise HTTPException(status_code=422, detail=f"Unknown payment method: {raw_method}")
    method_enum = PaymentMethod[enum_name]

    proof_path = None
    if proof_file and proof_file.filename:
        ext = os.path.splitext(proof_file.filename)[1]
        unique_name = f"proof_{registration_id}_{uuid.uuid4().hex[:8]}{ext}"
        proof_path = os.path.join(PAYMENT_RECEIPT_DIR, unique_name)
        with open(proof_path, "wb+") as f:
            f.write(await proof_file.read())

    auto_ref = f"REF-{uuid.uuid4().hex[:10].upper()}"

    existing = db.query(Payment).filter(
        Payment.registration_id == registration_id
    ).first()

    if existing:
        existing.payment_method = method_enum
        existing.payment_reference = auto_ref
        existing.payment_amount = payment_amount
        existing.payment_date = datetime.utcnow()
        if proof_path:
            registration.payment_proof = proof_path
        db.commit()
        try:
            ev = db.query(Event).filter(Event.id == event_id).first()
            if registration.user and registration.user.email:
                proof_of_payment_received_email(
                    recipient_email=registration.user.email,
                    firstname=registration.user.firstname or "Participant",
                    event_name=ev.event if ev else "the event",
                    background_tasks=background_tasks,
                    db=db,
                )
        except Exception:
            pass
        return {"message": "Payment details updated", "payment_id": existing.id}

    new_payment = Payment(
        registration_id=registration_id,
        payment_date=datetime.utcnow(),
        payment_method=method_enum,
        payment_reference=auto_ref,
        payment_amount=payment_amount,
        payment_status=PaymentStatus.PENDING,
    )
    db.add(new_payment)
    if proof_path:
        registration.payment_proof = proof_path
    db.commit()
    try:
        ev = db.query(Event).filter(Event.id == event_id).first()
        if registration.user and registration.user.email:
            proof_of_payment_received_email(
                recipient_email=registration.user.email,
                firstname=registration.user.firstname or "Participant",
                event_name=ev.event if ev else "the event",
                background_tasks=background_tasks,
                db=db,
            )
    except Exception:
        pass
    return {"message": "Payment submitted successfully", "payment_id": new_payment.id}


@router.post("/register-with-payment/")
async def register_with_payment(
    user_id: Optional[int] = Form(None),
    event_id: int = Form(...),
    participation_role: str = Form(...),
    payment_method: str = Form(...),
    payment_amount: float = Form(...),
    proof_file: UploadFile = File(...),
    # New-user fields — only sent when registering without an existing account
    new_firstname: Optional[str] = Form(None),
    new_lastname: Optional[str] = Form(None),
    new_email: Optional[str] = Form(None),
    new_phone: Optional[str] = Form(None),
    new_title: Optional[str] = Form(None),
    new_middle_name: Optional[str] = Form(None),
    new_country_id: Optional[int] = Form(None),
    new_profession: Optional[str] = Form(None),
    new_gender: Optional[str] = Form(None),
    new_organisation: Optional[str] = Form(None),
    new_position: Optional[str] = Form(None),
    background_tasks: BackgroundTasks = None,
    db: Session = Depends(get_db),
):
    """Create registration + payment record together, only after proof is uploaded.
    When user_id is None, creates the user account atomically with the registration.
    """
    from models.models import (
        PaymentMethod, PaymentStatus, ParticipationRole,
        UserProfile, UserRole, AccountVerification, Role,
    )
    import utils.mailer_util as mailer_util

    # ── Create new user account when no user_id provided ─────────────────────
    # Credentials are generated here only to satisfy the required password column;
    # they are never emailed at this point. verify_payment() issues and emails a
    # fresh password once the secretariat verifies the payment.
    if not user_id:
        if not new_email:
            raise HTTPException(status_code=422, detail="Email is required for new registrations.")
        existing_account = db.query(User).filter(User.email == new_email).first()
        if existing_account:
            user_id = existing_account.id
        else:
            if not new_firstname or not new_lastname:
                raise HTTPException(status_code=422, detail="Name is required for new registrations.")
            auth_dep = Auth(db)
            hashed = auth_dep.hash_password(auth_dep.generate_random_password())

            new_user = User(
                firstname=new_firstname,
                lastname=new_lastname,
                phone=new_phone,
                email=new_email,
                hashed_password=hashed,
                verified=False,
                must_change_password=True,
            )
            db.add(new_user)
            db.flush()
            user_id = new_user.id

            role = db.query(Role).filter(Role.role == "User").first()
            if role:
                db.add(UserRole(user_id=user_id, role_id=role.id))

            if new_country_id:
                db.add(UserProfile(
                    user_id=user_id,
                    title=new_title or "",
                    middle_name=new_middle_name or "",
                    country_id=new_country_id,
                    profession=new_profession or "",
                    gender=new_gender or "",
                    organisation=new_organisation or "",
                    position=new_position or "",
                ))

            db.add(AccountVerification(
                user_id=user_id,
                verification_token=str(uuid.uuid4()),
                expires_at=datetime.utcnow() + timedelta(hours=1),
            ))
            db.flush()
    # ─────────────────────────────────────────────────────────────────────────

    existing = db.query(Registration).filter(
        Registration.user_id == user_id,
        Registration.event_id == event_id,
        Registration.deleted_at == None,
    ).first()

    # Resolve enums
    raw_method = payment_method.strip()
    enum_name = METHOD_MAP.get(raw_method.lower())
    if not enum_name:
        raise HTTPException(status_code=422, detail=f"Unknown payment method: {raw_method}")
    method_enum = PaymentMethod[enum_name]

    # Save proof file
    if not proof_file or not proof_file.filename:
        raise HTTPException(status_code=422, detail="Proof of payment file is required.")
    ext = os.path.splitext(proof_file.filename)[1]
    unique_name = f"proof_new_{user_id}_{event_id}_{uuid.uuid4().hex[:8]}{ext}"
    proof_path = os.path.join(PAYMENT_RECEIPT_DIR, unique_name)
    with open(proof_path, "wb+") as f:
        f.write(await proof_file.read())

    if existing:
        # Already registered — update proof and reset payment to pending re-verification
        try:
            existing.payment_proof = proof_path
            existing.paid = False
            existing_payment = db.query(Payment).filter(Payment.registration_id == existing.id).first()
            if existing_payment:
                existing_payment.payment_method = method_enum
                existing_payment.payment_amount = payment_amount
                existing_payment.payment_status = PaymentStatus.PENDING
                existing_payment.payment_date = datetime.utcnow()
            else:
                db.add(Payment(
                    registration_id=existing.id,
                    payment_date=datetime.utcnow(),
                    payment_method=method_enum,
                    payment_reference=f"REF-{uuid.uuid4().hex[:10].upper()}",
                    payment_amount=payment_amount,
                    payment_status=PaymentStatus.PENDING,
                ))
            db.commit()
            # Credentials are only ever emailed at payment verification time (verify_payment).
            # At upload time we just confirm receipt of the proof.
            try:
                ev = db.query(Event).filter(Event.id == event_id).first()
                _u = db.query(User).filter(User.id == user_id).first()
                if _u and _u.email:
                    mailer_util.proof_of_payment_received_email(
                        recipient_email=_u.email,
                        firstname=_u.firstname or "Participant",
                        event_name=ev.event if ev else "the event",
                        background_tasks=background_tasks,
                        db=db,
                    )
            except Exception:
                pass
            return {"message": "Payment proof updated successfully", "registration_id": existing.id}
        except HTTPException:
            raise
        except Exception as e:
            db.rollback()
            if os.path.exists(proof_path):
                os.remove(proof_path)
            raise HTTPException(status_code=500, detail="Failed to update proof. Please try again.") from e

    try:
        role_enum = ParticipationRole[participation_role]
    except KeyError:
        raise HTTPException(status_code=422, detail=f"Unknown participation role: {participation_role}")

    try:
        new_registration = Registration(
            user_id=user_id,
            event_id=event_id,
            participation_role=role_enum,
            payment_proof=proof_path,
        )
        db.add(new_registration)
        db.flush()

        db.add(Payment(
            registration_id=new_registration.id,
            payment_date=datetime.utcnow(),
            payment_method=method_enum,
            payment_reference=f"REF-{uuid.uuid4().hex[:10].upper()}",
            payment_amount=payment_amount,
            payment_status=PaymentStatus.PENDING,
        ))
        db.commit()

        # Credentials are only ever emailed at payment verification time (verify_payment).
        # At upload time — whether the account is brand new or pre-existing — we just
        # confirm receipt of the proof.
        try:
            ev = db.query(Event).filter(Event.id == event_id).first()
            event_name = ev.event if ev else "the event"
            _user = db.query(User).filter(User.id == user_id).first()
            if _user and _user.email:
                mailer_util.proof_of_payment_received_email(
                    recipient_email=_user.email,
                    firstname=_user.firstname or "Participant",
                    event_name=event_name,
                    background_tasks=background_tasks,
                    db=db,
                )
        except Exception:
            pass

        return {"message": "Registration and payment proof submitted successfully", "registration_id": new_registration.id}
    except Exception as e:
        db.rollback()
        if os.path.exists(proof_path):
            os.remove(proof_path)
        raise HTTPException(status_code=500, detail="Failed to complete registration. Please try again.") from e


@router.post("/registration/{user_id}")
async def event_registration(
    request: Request,
    user_id: int,
    registration_schema: RegistrationSchema,
    # current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    # auth_dependency: Auth = Depends(get_auth_dependency),
):
    # user = db.query(User).filter(User.id == user_id).first()
    # client_ip = dependency.request_ip(request)

    # dependency.log_activity(
    #     1,
    #     "EVENT_REGISTRATION",
    #     user.email,
    #     client_ip,
    #     f"Event ID: {registration_schema.event_id}",
    # )

    # Check for existing registration
    existing_registration = (
        db.query(Registration)
        .filter(
            Registration.user_id == user_id,
            Registration.event_id == registration_schema.event_id,
        )
        .first()
    )

    if existing_registration:
        # Update existing registration
        existing_registration.participation_role = (
            registration_schema.participation_role
        )
        db.commit()
        db.refresh(existing_registration)
        return {
            "message": "Registration updated successfully",
            "registration_id": existing_registration.id,
        }

    # Create new registration
    new_registration = Registration(
        user_id=user_id,
        event_id=registration_schema.event_id,
        participation_role=registration_schema.participation_role,
    )
    db.add(new_registration)
    db.commit()
    db.refresh(new_registration)

    return {
        "message": "Registration successful",
        "registration_id": new_registration.id,
    }


@router.delete("/registration/{event_id}")
async def event_deregistration(
    request: Request,
    event_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    client_ip = dependency.request_ip(request)

    dependency.log_activity(
        1,
        "EVENT_DEREGISTER",
        "None",
        client_ip,
        f"Deregister event id {event_id}",
    )

    existing_registration = (
        db.query(Registration)
        .filter(
            Registration.user_id == current_user["user_id"],
            Registration.event_id == event_id,
        )
        .first()
    )

    try:
        db.query(Registration).filter(
            Registration.id == existing_registration.id
        ).delete(synchronize_session=False)
        db.commit()
    except Exception as error:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete user event registration",
        ) from error


PAYMENT_RECEIPT_DIR = "uploads/payment_receipts"
if not os.path.exists(PAYMENT_RECEIPT_DIR):
    os.makedirs(PAYMENT_RECEIPT_DIR)


@router.delete("/deregister_participant/{registration_id}")
async def admin_deregister_participant(
    request: Request,
    registration_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Admin-only: deregister any participant by registration ID."""
    auth_dependency.secure_access("ADMIN_DASHBOARD", current_user["user_id"])

    registration = db.query(Registration).filter(Registration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    try:
        db.query(Registration).filter(
            Registration.id == registration.id
        ).delete(synchronize_session=False)
        db.commit()
    except Exception as error:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to deregister participant",
        ) from error

    client_ip = dependency.request_ip(request)
    dependency.log_activity(
        current_user["user_id"],
        "ADMIN_DEREGISTER",
        current_user["username"],
        client_ip,
        f"Admin deregistered registration ID {registration_id}",
    )
    return {"detail": "Participant successfully deregistered"}


@router.post("/upload_payment_proof/{event_id}")
async def upload_payment_proof(
    event_id: int,
    current_user: user_dependency,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """User uploads proof of payment for their registration."""
    registration = db.query(Registration).filter(
        Registration.user_id == current_user["user_id"],
        Registration.event_id == event_id,
        Registration.deleted_at == None,
    ).first()

    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    try:
        ext = os.path.splitext(file.filename)[1]
        unique_name = f"proof_{registration.id}_{uuid.uuid4().hex[:8]}{ext}"
        file_path = os.path.join(PAYMENT_RECEIPT_DIR, unique_name)
        with open(file_path, "wb+") as f:
            f.write(await file.read())

        registration.payment_proof = file_path
        db.commit()

        return JSONResponse(content={"status": "success", "payment_proof": file_path})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/verify_payment/{registration_id}")
async def verify_payment(
    request: Request,
    registration_id: int,
    current_user: user_dependency,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Admin verifies a participant's payment, setting paid=True."""
    auth_dependency.secure_access(["ADMIN_DASHBOARD", "VERIFY_PAYMENT"], current_user["user_id"])

    registration = db.query(Registration).filter(Registration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    registration.paid = True

    # Generate a fresh password so credentials are always included in the verified email
    new_plain_password = None
    try:
        new_plain_password = auth_dependency.generate_random_password()
        registration.user.hashed_password = auth_dependency.hash_password(new_plain_password)
        registration.user.must_change_password = True
    except Exception as _pe:
        logger.error(f"Failed to generate password for verified user: {_pe}")

    db.commit()

    try:
        from utils.mailer_util import payment_verified_email
        event = db.query(Event).filter(Event.id == registration.event_id).first()
        if registration.user and registration.user.email and event:
            payment_verified_email(
                recipient_email=registration.user.email,
                firstname=registration.user.firstname or "Participant",
                event_name=event.event,
                password=new_plain_password,
                background_tasks=background_tasks,
                db=db,
                sent_by_user_id=current_user["user_id"],
            )
    except Exception as _e:
        logger.error(f"Failed to send payment verified email: {_e}")

    client_ip = dependency.request_ip(request)
    dependency.log_activity(
        current_user["user_id"],
        "VERIFY_PAYMENT",
        current_user["username"],
        client_ip,
        f"Verified payment for registration ID {registration_id}",
    )
    return {"detail": "Payment verified successfully"}


@router.put("/unverify_payment/{registration_id}")
async def unverify_payment(
    request: Request,
    registration_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Admin marks a participant's payment as unpaid."""
    auth_dependency.secure_access(["ADMIN_DASHBOARD", "VERIFY_PAYMENT"], current_user["user_id"])

    registration = db.query(Registration).filter(Registration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    registration.paid = False
    db.commit()

    client_ip = dependency.request_ip(request)
    dependency.log_activity(
        current_user["user_id"],
        "UNVERIFY_PAYMENT",
        current_user["username"],
        client_ip,
        f"Marked payment as unpaid for registration ID {registration_id}",
    )
    return {"detail": "Payment marked as unpaid"}


@router.post("/upload_document/")
async def upload_document(
    user: user_dependency,
    file: UploadFile = File(...),
    file_name: str = Form(...),
    doc_type: str = Form(...),
    access_level: str = Form(...),
    event_id: int = Form(...),
    db: Session = Depends(get_db),
):
    try:
        unique_dir = os.path.join(
            EVENT_DOCUMENT_DIR,
            f"{event_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}",
        )
        os.makedirs(unique_dir, exist_ok=True)
        file_path = os.path.join(unique_dir, file.filename)
        with open(file_path, "wb+") as file_object:
            file_object.write(await file.read())

        event_document_model = Document(
            event_id=event_id,
            document_type=doc_type,
            file_type=file.content_type,
            file_name=file.filename,
            name=file_name,
            path=file_path,
            access_level=access_level,
        )
        db.add(event_document_model)
        db.commit()
        db.refresh(event_document_model)

        return JSONResponse(
            content={
                "status": "success",
                "message": f"File '{file.filename}' uploaded to '{unique_dir}'",
                "file_path": file_path,
            },
            status_code=200,
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")


@router.post("/upload_banner/{event_id}")
async def upload_banner(
    event_id: int,
    user: user_dependency,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    try:
        ext = os.path.splitext(file.filename)[1]
        unique_name = f"{event_id}_{uuid.uuid4().hex[:8]}{ext}"
        file_path = os.path.join(EVENT_BANNER_DIR, unique_name)
        with open(file_path, "wb+") as f:
            f.write(await file.read())

        event = db.query(Event).filter(Event.id == event_id).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        event.banner_image = file_path
        db.commit()

        return JSONResponse(content={"status": "success", "banner_image": file_path})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/delete_document/{document_id}")
async def delete_document(
    request: Request,
    document_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    auth_dependency.secure_access("DELETE_EVENT", current_user["user_id"])

    document = get_object(document_id, db, Document)

    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    document_folder = os.path.dirname(document.path)
    if os.path.exists(document_folder):
        try:
            shutil.rmtree(document_folder)
        except Exception as e:
            raise HTTPException(
                status_code=500, detail=f"Failed to delete folder: {str(e)}"
            )
    else:
        raise HTTPException(status_code=404, detail="Folder not found on disk")

    db.delete(document)
    db.commit()

    return {"status": "success", "message": f"Document and folder deleted successfully"}


@router.post("/add_link/")
async def add_link(
    request: Request,
    link_schema: LinkSchema,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    client_ip = dependency.request_ip(request)
    dependency.log_activity(
        current_user["user_id"],
        "ADD_EVENT",
        current_user["username"],
        client_ip,
        link_schema.event_id,
    )

    create_event_link_model = Link(
        event_id=link_schema.event_id,
        name=link_schema.name,
        link=str(link_schema.link),
    )

    db.add(create_event_link_model)
    db.commit()
    return link_schema


@router.put("/update_link/{link_id}")
async def update_link(
    request: Request,
    link_id: int,
    current_user: user_dependency,
    link_schema: LinkSchema,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    auth_dependency.secure_access("UPDATE_EVENT", current_user["user_id"])

    client_ip = dependency.request_ip(request)

    dependency.log_activity(
        current_user["user_id"],
        "UPDATE_EVENT",
        current_user["username"],
        client_ip,
        f"Update event id {link_id}",
    )
    link_model = get_object(link_id, db, Link)

    link_model.name = link_schema.name
    link_model.link = str(link_schema.link)

    db.commit()
    db.refresh(link_model)
    return link_schema


@router.delete("/delete_link/{link_id}")
async def delete_link(
    request: Request,
    link_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    dependency: Dependency = Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    auth_dependency.secure_access("DELETE_EVENT", current_user["user_id"])

    dependency.hard_delete(Link, link_id)

    client_ip = dependency.request_ip(request)
    link = get_object(link_id, db, Link)

    client_ip = dependency.request_ip(request)
    dependency.log_activity(
        current_user["user_id"],
        "DELETE_EVENT",
        current_user["username"],
        client_ip,
        f"Delete event id {link} event {link.name}",
    )
    return {"detail": "link Successfully deleted"}


@router.get("/{event_id}/participants/download")
async def download_event_participants(
    request: Request,
    event_id: int,
    current_user: user_dependency,
    paid: Literal["all", "true", "false"] = Query("all"),
    role_category: Literal["all", "secretariat", "djcc", "local_secretariat", "other"] = Query("all"),
    db: Session = Depends(get_db),
    dependency=Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    # Permission check
    auth_dependency.secure_access("VIEW_EVENT", current_user["user_id"])
    client_ip = dependency.request_ip(request)

    dependency.log_activity(
        current_user["user_id"],
        "DOWNLOAD_PARTICIPANTS",
        current_user["username"],
        client_ip,
        f"Downloaded participants for event {event_id} with filter paid={paid}, role_category={role_category}",
    )

    # Eager-load registrations + user/profile/country in bulk (see get_event()
    # / download_participant_badges_pdf for the same fix and why it matters).
    event = (
        db.query(Event)
        .options(
            selectinload(Event.registrations)
                .selectinload(Registration.user)
                .selectinload(User.user_profile)
                .joinedload(UserProfile.country),
        )
        .filter(Event.id == event_id, Event.deleted_at == None)
        .first()
    )
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    # Collect participant details
    participants = []
    for reg in event.registrations:
        user = reg.user
        profile = user.user_profile[0] if user.user_profile else None
        # badge_* fields (set by admin bulk-import/add) take priority over a
        # general profile when both exist — see get_event() for the same
        # convention. Also lets badge-only imports (no UserProfile at all,
        # e.g. Local Secretariat) show their country/organisation/position
        # here instead of coming back blank.
        country = reg.badge_country or (profile.country.country if profile and profile.country else None)
        organisation = reg.badge_organisation or (profile.organisation if profile else None)

        # Convert participation_role to string key (adjust this if ParticipationRole is Enum)
        role_key = (
            reg.participation_role.name
            if hasattr(reg.participation_role, "name")
            else str(reg.participation_role).lower()
        )

        participants.append(
            {
                "registration_id": reg.id,
                "firstname": user.firstname,
                "lastname": user.lastname,
                "email": user.email,
                "phone": user.phone,
                "title": reg.badge_prefix or (profile.title if profile else ""),
                "middle_name": profile.middle_name if profile else "",
                "gender": profile.gender if profile else "",
                "position": reg.badge_position or (profile.position if profile else ""),
                "organisation": organisation,
                "country": country,
                "participation_role": PARTICIPATION_ROLE_MAP.get(role_key, role_key),
                "role_key": role_key,
                "paid": reg.paid,
                "has_proof": bool(reg.payment_proof),
                "registered_at": reg.registered_at,
            }
        )

    # Filter by paid status — "Paid" means paid OR proof of payment uploaded (POP),
    # matching the confirmed-participants set shown on the Participants tab.
    if paid != "all":
        if paid == "true":
            participants = [p for p in participants if p["paid"] or p["has_proof"]]
        else:
            participants = [p for p in participants if not p["paid"] and not p["has_proof"]]

    # Filter by role category. Secretariat and DJCC members are only included
    # when explicitly selected via their own filter pill — the general "All"
    # export excludes them since they're downloaded separately.
    if role_category == "local_secretariat":
        participants = [p for p in participants if p["role_key"] in LOCAL_SECRETARIAT_ROLES]
    elif role_category in ("secretariat", "djcc"):
        participants = [p for p in participants if p["role_key"] == role_category]
    else:
        participants = [
            p for p in participants
            if p["role_key"] not in ({"secretariat", "djcc"} | LOCAL_SECRETARIAT_ROLES)
        ]

    if not participants:
        raise HTTPException(status_code=404, detail="No participants found")

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"

    headers = [
        "ID",
        "Title",
        "First Name",
        "Middle Name",
        "Last Name",
        "Gender",
        "Email",
        "Phone",
        "Organisation",
        "Position",
        "Country",
        "Participation Role",
        "Paid",
        "Registered At",
    ]
    ws.append(headers)

    for p in participants:
        ws.append(
            [
                p["registration_id"],
                p["title"],
                p["firstname"],
                p["middle_name"],
                p["lastname"],
                p["gender"],
                p["email"],
                p["phone"],
                p["organisation"] or "",
                p["position"],
                p["country"] or "",
                p["participation_role"],
                "Yes" if p["paid"] else ("Pending (Proof Uploaded)" if p["has_proof"] else "No"),
                p["registered_at"].strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )

    # Stream response
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    safe_event_name = sanitize_filename(event.event)
    ascii_filename = f"{safe_event_name}_participants.xlsx"
    utf8_filename = urllib.parse.quote(ascii_filename)

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={ascii_filename}; filename*=UTF-8''{utf8_filename}"
        },
    )


@router.get("/with-registration/{user_id}")
def list_events_with_user_registration(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),  # or your custom dependency
):
    # Get all events
    events = db.query(Event).all()

    # Get all registrations by the user
    user_regs = (
        db.query(Registration)
        .filter(Registration.user_id == user_id, Registration.deleted_at.is_(None))
        .all()
    )

    # Build a lookup dictionary for quick access: event_id -> registration
    reg_map = {reg.event_id: reg for reg in user_regs}

    # Construct result
    result = []
    for event in events:
        registration = reg_map.get(event.id)
        result.append(
            {
                "event": {
                    "id": event.id,
                    "title": event.event,
                    "theme": event.theme,
                    "description": event.description,
                    "start_date": event.start_date,
                    "end_date": event.end_date,
                    "location": event.location,
                    "country": event.country.country if event.country else None,
                    "org_unit": event.org_unit.name if event.org_unit else None,
                },
                "registered": bool(registration),
                "registration_details": (
                    {
                        "registration_id": registration.id,
                        "participation_role": registration.participation_role,
                        "paid": registration.paid,
                        "payment_status": "Paid" if registration.paid else "Not Paid",
                        "registered_at": registration.registered_at,
                    }
                    if registration
                    else None
                ),
            }
        )

    return result


def hex_to_rgb(hex_color: str):
    """Convert a hex color string like '#a02626' to a 0.0-1.0 RGB tuple."""
    hex_color = hex_color.lstrip("#")
    return tuple(int(hex_color[i: i + 2], 16) / 255.0 for i in (0, 2, 4))


_STATIC_BADGE_FORM_NAME = "ecsa_badge_static_v1"


def _register_static_badge_form(c, logo_left, logo_right):
    """Register (once per canvas) a Form XObject holding the badge's static
    background/logos/flags, so a multi-badge PDF can stamp them via doForm()
    instead of re-encoding identical image bytes on every single page.

    Profiling a 264-badge export showed ~75% of render time going into
    reportlab's drawImage (MD5 digest + zlib compress + base85 encode),
    which reportlab redoes on every call even when it's literally the same
    ImageReader object as the previous badge — the 2 logos + 11 flags never
    change between badges. A Form XObject embeds that artwork once in the
    PDF and every badge just references it, cutting a 264-badge export from
    ~28s to a few seconds.
    """
    if getattr(c, "_ecsa_badge_static_registered", False):
        return
    W_mm, H_mm = 105.0, 148.0
    W, H = W_mm * mm, H_mm * mm

    def fy(y_from_top):
        return (H_mm - y_from_top) * mm

    c.beginForm(_STATIC_BADGE_FORM_NAME, lowerx=0, lowery=0, upperx=W, uppery=H)

    bg_path = "assets/badge_bg.jpg"
    if os.path.exists(bg_path):
        try:
            c.drawImage(convert_png_to_rgb(bg_path), 0, 0, W, H)
        except Exception:
            c.setFillColorRGB(0.82, 0.80, 0.78)
            c.rect(0, 0, W, H, fill=True, stroke=False)
    else:
        c.setFillColorRGB(0.82, 0.80, 0.78)
        c.rect(0, 0, W, H, fill=True, stroke=False)

    c.saveState()
    margin_mm = 4.0
    badge_scale = min((W_mm - 2 * margin_mm) / W_mm, (H_mm - 2 * margin_mm) / H_mm)
    c.translate((W - W * badge_scale) / 2, (H - H * badge_scale) / 2)
    c.scale(badge_scale, badge_scale)

    logo_h_mm = 14.0
    if logo_left:
        try:
            c.drawImage(logo_left, 5*mm, fy(3 + logo_h_mm),
                        width=27*mm, height=logo_h_mm*mm, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass
    if logo_right:
        try:
            c.drawImage(logo_right, 58*mm, fy(3 + logo_h_mm),
                        width=40*mm, height=logo_h_mm*mm, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

    qr_mm, qr_top = 22, 115.5
    flag_map  = _get_flag_images()
    all_flags = [flag_map.get(code) for code in ECSA_FLAG_CODES]
    valid     = [img for img in all_flags if img is not None]
    if valid:
        max_row_w_mm = 92.0
        n            = len(valid)
        f_gap_mm     = 0.9
        flag_w_mm    = min(7.0, (max_row_w_mm - (n - 1) * f_gap_mm) / n)
        flag_h_mm    = flag_w_mm / 1.545
        row_w_mm     = n * flag_w_mm + (n - 1) * f_gap_mm
        sx           = (W_mm - row_w_mm) / 2
        row_top      = qr_top + qr_mm + 5.5
        flag_rl_y    = fy(row_top + flag_h_mm)
        for j, img in enumerate(valid):
            try:
                c.drawImage(img,
                            (sx + j * (flag_w_mm + f_gap_mm)) * mm,
                            flag_rl_y,
                            flag_w_mm * mm, flag_h_mm * mm,
                            preserveAspectRatio=True)
            except Exception:
                pass

    c.restoreState()
    c.endForm()
    c._ecsa_badge_static_registered = True


def _render_badge_page(c, p, logo_left, logo_right, primary_rgb=None, secondary_rgb=None, blank=False):
    """Render one A6 ECSA conference badge matching the official name-tag design exactly.

    All positions derived from official ECSA-HC A6 (105×148 mm) badge PDFs.
    PyMuPDF extracted coordinates are in mm from top; converted to ReportLab
    (y from bottom) via fy(y_top_mm) = (148 - y_top_mm) * mm.

    blank=True is for generic role badges with no real registration behind
    them (empty Name/Designation/Organization for on-site handwriting): the
    QR still renders for visual consistency but encodes the role label as
    plain text instead of an attendance-confirmation URL, so scanning it
    doesn't try to open a page.
    """
    W_mm, H_mm = 105.0, 148.0
    W, H = W_mm * mm, H_mm * mm

    def fy(y_from_top):
        """mm-from-top  →  ReportLab points-from-bottom."""
        return (H_mm - y_from_top) * mm

    # ── Role colour ──────────────────────────────────────────────────────────
    role_raw   = str(p.get("participation_role_raw", "delegate")).lower()
    role_hex   = BADGE_ROLE_COLORS.get(role_raw, "#0095B6")
    role_rgb   = hex_to_rgb(role_hex)
    role_light = tuple(0.80 + v * 0.20 for v in role_rgb)
    banner_txt = (0.0, 0.0, 0.0) if role_hex == "#FFD700" else (1.0, 1.0, 1.0)
    role_label = BADGE_ROLE_LABELS.get(role_raw, role_raw.upper())

    # ── Background + logos + flags (static across every badge) ────────────────
    _register_static_badge_form(c, logo_left, logo_right)
    c.doForm(_STATIC_BADGE_FORM_NAME)

    # ── Inset all foreground content by a safety margin ────────────────────────
    # Plastic A6 card holders typically clip a few mm around the edge. The
    # background texture above stays full-bleed; everything drawn from here on
    # (text, QR code) is shrunk and centred so none of it gets cropped once
    # the badge is in its holder — matches the same transform baked into the
    # static form above so text lines up with the logos/flags exactly.
    c.saveState()
    margin_mm = 4.0
    badge_scale = min((W_mm - 2 * margin_mm) / W_mm, (H_mm - 2 * margin_mm) / H_mm)
    c.translate((W - W * badge_scale) / 2, (H - H * badge_scale) / 2)
    c.scale(badge_scale, badge_scale)

    # ── Event title ───────────────────────────────────────────────────────────
    # Extracted: title1 baseline y=29.3 mm (21.3 pt), title2 baseline y=38.1 mm (17.8 pt)
    raw_name = normalize_event_name(str(p.get("event_name") or ""))
    parts = raw_name.split(" & ", 1)
    if len(parts) == 2:
        title1 = parts[0].strip()
        title2 = "& " + parts[1].strip()
    else:
        words = raw_name.split()
        mid   = max(1, (len(words) + 1) // 2)
        title1 = " ".join(words[:mid])
        title2 = " ".join(words[mid:])

    # Shifted up from the original extracted positions to tighten the gap
    # below the logos (and, further down, to make room for larger info-row
    # text without pushing the QR code or anything below it).
    top_shift = 4.5

    c.setFillColorRGB(0.969, 0.580, 0.114)          # ECSA orange
    c.setFont("Helvetica-Bold", 21)
    c.drawCentredString(W / 2, fy(29.3 - top_shift + 21 * 25.4/72 * 0.75), title1)

    c.setFillColorRGB(0.06, 0.06, 0.06)             # near-black (matches official)
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(W / 2, fy(38.1 - top_shift + 18 * 25.4/72 * 0.75), title2)

    # ── Dates & location ──────────────────────────────────────────────────────
    # Extracted: date baseline y=46.1 mm, 11 pt
    dates    = str(p.get("event_dates") or "")
    loc      = str(p.get("location") or "")
    date_str = f"{dates}  |  {loc}" if (dates and loc) else (dates or loc)
    c.setFillColorRGB(0.1, 0.1, 0.1)
    # Auto-size date string to fit within usable width (94 mm)
    for date_fs in (11, 9, 8, 7):
        c.setFont("Helvetica-Bold", date_fs)
        if c.stringWidth(date_str, "Helvetica-Bold", date_fs) <= 94 * mm:
            break
    c.drawCentredString(W / 2, fy(46.1 - top_shift + date_fs * 25.4/72 * 0.75), date_str)

    # ── THEME box ─────────────────────────────────────────────────────────────
    # Extracted: rect (9.9, 53.9)→(28.1, 59.4) mm  |  label 12.6 pt
    c.setFillColorRGB(0.0, 0.681, 0.938)
    c.rect(9.9*mm, fy(59.4 - top_shift), 18.2*mm, 5.5*mm, fill=True, stroke=False)
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 11)
    theme_box_cy = (53.9 + 59.4) / 2 - top_shift    # 56.65 mm from top, shifted
    c.drawCentredString(19.0*mm, fy(theme_box_cy + 11 * 25.4/72 * 0.3), "THEME:")

    # ── Theme text ────────────────────────────────────────────────────────────
    # Extracted: line1 y=60.7 mm, line2 y=64.6 mm  |  9 pt
    theme = str(p.get("event_theme") or "")
    c.setFillColorRGB(0.06, 0.06, 0.06)
    c.setFont("Helvetica-Bold", 8.5)
    if theme:
        limit = 55
        if len(theme) <= limit:
            c.drawString(10.2*mm, fy(60.7 - top_shift + 8.5 * 25.4/72 * 0.75), theme)
        else:
            bp = theme.rfind(" ", 0, limit) or limit
            c.drawString(10.2*mm, fy(60.7 - top_shift + 8.5 * 25.4/72 * 0.75), theme[:bp].rstrip())
            c.drawString(10.2*mm, fy(64.6 - top_shift + 8.5 * 25.4/72 * 0.75), theme[bp:].strip())

    def _shrink_to_fit(text, start_size, floor_size, max_width_pt, font="Helvetica-Bold"):
        """Step the font size down (in 0.5pt steps) until text fits max_width_pt,
        or floor_size is reached — for role labels/values too long to fit as-is."""
        size = start_size
        while size > floor_size and c.stringWidth(text, font, size) > max_width_pt:
            size -= 0.5
        return size

    def _fit_text_and_size(text, start_size, floor_size, max_width_pt, font="Helvetica-Bold"):
        """Shrink font first; if even the floor size still overflows (extremely
        long values), fall back to truncating with an ellipsis so text never
        runs past its box/the page edge."""
        size = _shrink_to_fit(text, start_size, floor_size, max_width_pt, font)
        if c.stringWidth(text, font, size) <= max_width_pt:
            return text, size
        truncated = text
        while truncated and c.stringWidth(truncated + "…", font, size) > max_width_pt:
            truncated = truncated[:-1]
        return (truncated.rstrip() + "…"), size

    def _smart_shorten_org(text, start_size, floor_size, max_width_pt, font="Helvetica-Bold"):
        """Organization names that need shrinking to fit shouldn't just render
        tiny — try the org's own acronym if it's spelled out in parentheses
        (e.g. "...Institute for Policy and Research (SAIPAR)" → "SAIPAR"),
        generic word abbreviation (Institute→Inst., University→Univ., ...),
        or both, and use whichever form lets the text stay at the LARGEST
        font size (not just whichever fits at the smallest). The original is
        kept whenever it already fits at the starting size — abbreviation
        only kicks in once shrinking would otherwise be needed. Falls back to
        an ellipsis only if nothing fits even at the floor size."""
        def fits(t, size):
            return c.stringWidth(t, font, size) <= max_width_pt

        candidates = [text]
        m = re.search(r"\(([A-Z]{2,12})\)\s*,?\s*(.*)$", text)
        acronym_form = None
        if m:
            acronym, rest = m.group(1), m.group(2).strip(" ,")
            acronym_form = f"{acronym}, {rest}" if rest else acronym
            candidates.append(acronym_form)
        abbreviated = abbreviate_org_words(text)
        if abbreviated != text:
            candidates.append(abbreviated)
        if acronym_form:
            abbreviated_acronym = abbreviate_org_words(acronym_form)
            if abbreviated_acronym != acronym_form:
                candidates.append(abbreviated_acronym)
        generated_acronym = generate_org_acronym(text)
        if generated_acronym and generated_acronym not in candidates:
            candidates.append(generated_acronym)

        best_cand, best_size, best_fits = text, floor_size, False
        best_score = -1
        for cand in candidates:
            cand_size = _shrink_to_fit(cand, start_size, floor_size, max_width_pt, font)
            cand_fits = fits(cand, cand_size)
            score = cand_size if cand_fits else -1
            if score > best_score:
                best_score, best_cand, best_size, best_fits = score, cand, cand_size, cand_fits

        if best_fits:
            return best_cand, best_size

        # Nothing fits even at floor size — ellipsis-truncate the shortest candidate.
        shortest = min(candidates, key=len)
        return _fit_text_and_size(shortest, start_size, floor_size, max_width_pt, font)

    # ── Role banner ───────────────────────────────────────────────────────────
    # Extracted: rect (7.8, 74.0)→(97.9, 86.3) mm  |  text 30 pt
    # Widened a couple mm on each side (was 7.8→97.9) to give long role labels
    # a bit more room before the font has to shrink.
    banner_x0, banner_x1 = 6.0, 99.0
    banner_top, banner_bottom = 74.0 - top_shift, 86.3 - top_shift
    c.setFillColorRGB(*role_rgb)
    c.rect(banner_x0*mm, fy(banner_bottom), (banner_x1 - banner_x0)*mm,
           (banner_bottom - banner_top) * mm, fill=True, stroke=False)

    fsize = max(18, min(30, int(330 // max(len(role_label), 1))))
    role_label, fsize = _fit_text_and_size(role_label, fsize, 14, (banner_x1 - banner_x0 - 4) * mm)
    banner_cy = (banner_top + banner_bottom) / 2
    c.setFont("Helvetica-Bold", fsize)
    c.setFillColorRGB(*banner_txt)
    c.drawCentredString(W / 2, fy(banner_cy + fsize * 25.4/72 * 0.3), role_label)

    # ── Info rows: Name / Designation / Organization ──────────────────────────
    # Taller rows with bigger text than the original extracted design (89.1
    # onward) — the top_shift above freed up just enough room that these
    # still end before the QR code without moving anything below them.
    # Both the label and content backgrounds are a couple mm wider than the
    # original extraction (was 7.8→97.8) for the same reason as the banner.
    title_raw = str(p.get("title") or "").strip().rstrip(".")
    firstname = str(p.get("firstname") or "").strip()
    lastname  = str(p.get("lastname")  or "").strip()
    if title_raw:
        full_name = f"{title_raw}. {firstname} {lastname}".strip()
    else:
        full_name = f"{firstname} {lastname}".strip()
    LBL_FS, VAL_FS = 12, 11
    ROW_X0, ROW_X1 = 6.0, 99.0
    # Label box widths were previously sized to an old fixed template and had
    # several mm of dead padding around the label word itself (e.g.
    # "Designation" only needs ~24mm but had a 34.5mm box). Narrowed to just
    # fit the label word + a little padding, so the freed width goes to the
    # value instead of sitting empty — lets long designations/orgs render at
    # a noticeably bigger font before they need to shrink.
    row_specs = [
        # (label, lbl_x0, lbl_x1, y0, y1, value)
        # Organization keeps a generous cap (not the tight 50-char one the
        # others use) so a trailing "(ACRONYM)" isn't chopped off before
        # _smart_shorten_org below gets a chance to use it.
        ("Name",         ROW_X0, 20.5,  84.6,  92.6, full_name[:55]),
        ("Designation",  ROW_X0, 33.5,  95.2, 103.2, str(p.get("position")     or "")[:50]),
        ("Organization", ROW_X0, 35.0, 105.8, 113.8, str(p.get("organisation") or "")[:200]),
    ]

    for lbl, lx0, lx1, y0f, y1f, val in row_specs:
        row_h_mm = y1f - y0f
        rl_bot   = fy(y1f)
        lbl_w_mm = lx1 - lx0
        cont_w_mm = ROW_X1 - lx1
        row_cy   = (y0f + y1f) / 2

        # Label (darker cyan)
        c.setFillColorRGB(0.0, 0.681, 0.938)
        c.rect(lx0*mm, rl_bot, lbl_w_mm*mm, row_h_mm*mm, fill=True, stroke=False)
        c.setFillColorRGB(1, 1, 1)
        lbl_fs = _shrink_to_fit(lbl, LBL_FS, 9, (lbl_w_mm - 2) * mm)
        c.setFont("Helvetica-Bold", lbl_fs)
        c.drawCentredString(((lx0 + lx1) / 2)*mm,
                            fy(row_cy + lbl_fs * 25.4/72 * 0.3), lbl)

        # Content (light tint) — font shrinks for long designations/orgs/names
        # instead of just being cut off; Organization additionally tries an
        # acronym / word-abbreviation shorten before resorting to ellipsis.
        c.setFillColorRGB(*role_light)
        c.rect(lx1*mm, rl_bot, cont_w_mm*mm, row_h_mm*mm, fill=True, stroke=False)
        c.setFillColorRGB(0.06, 0.06, 0.06)
        if lbl == "Organization":
            val, val_fs = _smart_shorten_org(val, VAL_FS, 7, (cont_w_mm - 3) * mm)
        else:
            val, val_fs = _fit_text_and_size(val, VAL_FS, 7, (cont_w_mm - 3) * mm)
        c.setFont("Helvetica-Bold", val_fs)
        c.drawString((lx1 + 1.5)*mm, fy(row_cy + val_fs * 25.4/72 * 0.3), val)

    # ── QR code ───────────────────────────────────────────────────────────────
    # qr_top stays fixed (keeps the gap to the Organization row above); the
    # code grows downward from there, so making it bigger doesn't need to
    # touch that gap.
    qr_mm   = 22
    qr_top  = 115.5                  # mm from top
    # blank badges keep the QR purely for visual/design consistency — there's
    # no real registration behind them, so it encodes plain text (the role
    # label) rather than a URL, so scanning it doesn't try to open a page.
    qr_payload = (
        f"{CLIENT_ORIGIN}/event-attendance/{p['event_id']}?reg={p['registration_id']}"
        if not blank else role_label
    )
    try:
        qr_img = qrcode.make(qr_payload)
        qr_buf = BytesIO()
        qr_img.save(qr_buf, format="PNG")
        qr_buf.seek(0)
        c.drawImage(ImageReader(qr_buf),
                    (W_mm - qr_mm) / 2 * mm, fy(qr_top + qr_mm),
                    qr_mm * mm, qr_mm * mm)
    except Exception:
        pass

    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.setFont("Helvetica", 6.5)
    c.drawCentredString(W / 2, fy(qr_top + qr_mm + 3.0),
                        "Scan QR code to confirm attendance" if not blank else "")

    # ECSA member-state flags are drawn once as part of the static form
    # registered at the top of this function (see _register_static_badge_form).

    c.restoreState()
    c.showPage()


@router.get("/{event_id}/participants/badges")
async def download_participant_badges_pdf(
    request: Request,
    event_id: int,
    current_user: user_dependency,
    paid: Literal["all", "true", "false"] = Query("all"),
    role_category: Literal["all", "secretariat", "djcc", "local_secretariat", "other"] = Query("all"),
    user_id: Optional[int] = Query(None),
    user_ids: Optional[str] = Query(None, description="Comma-separated user IDs to include"),
    db: Session = Depends(get_db),
    dependency=Depends(get_dependency),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    auth_dependency.secure_access("VIEW_EVENT", current_user["user_id"])
    client_ip = dependency.request_ip(request)
    dependency.log_activity(
        current_user["user_id"],
        "DOWNLOAD_BADGES",
        current_user["username"],
        client_ip,
        f"Downloaded participant badges for event {event_id} with filter paid={paid}, role_category={role_category}",
    )

    # Eager-load registrations + user/profile/country in bulk — same N+1 fix
    # as get_event(): lazily accessing reg.user / user.user_profile / profile
    # .country per registration issued ~3 extra round trips per person, which
    # dominated wall-clock time for a 264-badge export even after the PDF
    # rendering itself was optimized.
    event = (
        db.query(Event)
        .options(
            joinedload(Event.org_unit),
            selectinload(Event.registrations)
                .selectinload(Registration.user)
                .selectinload(User.user_profile)
                .joinedload(UserProfile.country),
        )
        .filter(Event.id == event_id, Event.deleted_at == None)
        .first()
    )
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    primary_color = (event.org_unit.primary_color or "#0095B6") if event.org_unit else "#0095B6"
    secondary_color = (event.org_unit.secondary_color or "#F7941D") if event.org_unit else "#F7941D"
    primary_rgb = hex_to_rgb(primary_color)
    secondary_rgb = hex_to_rgb(secondary_color)

    participants = []
    for reg in event.registrations:
        user = reg.user
        profile = user.user_profile[0] if user.user_profile else None
        # badge_* fields (set by admin bulk-import/add) take priority over a
        # general profile when both exist — e.g. someone re-imported for a
        # different event with fresh title/org data shouldn't have it
        # silently overridden by an older, unrelated profile.
        country = reg.badge_country or (profile.country.country if profile and profile.country else None)
        organisation = reg.badge_organisation or (profile.organisation if profile else None)
        role_key = (
            reg.participation_role.name
            if hasattr(reg.participation_role, "name")
            else str(reg.participation_role).lower()
        )
        participants.append(
            {
                "registration_id": reg.id,
                "user_id": user.id,
                "event_id": event_id,
                "title": (reg.badge_prefix or (profile.title if profile else "")),
                "firstname": user.firstname,
                "middle_name": profile.middle_name if profile else "",
                "lastname": user.lastname,
                "position": (reg.badge_position or (profile.position if profile else "")),
                "organisation": organisation,
                "country": country,
                "participation_role": PARTICIPATION_ROLE_MAP.get(role_key, role_key),
                "event_name": event.event,
                "location": event.location or "",
                "event_theme": event.theme or "",
                "event_dates": _fmt_event_dates(event),
                "participation_role_raw": role_key,
                "paid": reg.paid,
                "has_proof": bool(reg.payment_proof),
            }
        )

    # "Paid" means paid OR proof of payment uploaded (POP), matching the
    # confirmed-participants set shown on the Participants tab.
    if paid != "all":
        if paid == "true":
            participants = [p for p in participants if p["paid"] or p["has_proof"]]
        else:
            participants = [p for p in participants if not p["paid"] and not p["has_proof"]]

    if role_category != "all":
        if role_category == "other":
            participants = [
                p for p in participants
                if p["participation_role_raw"] not in ({"secretariat", "djcc"} | LOCAL_SECRETARIAT_ROLES)
            ]
        elif role_category == "local_secretariat":
            participants = [p for p in participants if p["participation_role_raw"] in LOCAL_SECRETARIAT_ROLES]
        else:
            participants = [p for p in participants if p["participation_role_raw"] == role_category]

    # Optional single-participant filter (used by badge preview download button)
    if user_id is not None:
        participants = [p for p in participants if p.get("user_id") == user_id]

    if user_ids:
        selected_ids = {int(uid) for uid in user_ids.split(",") if uid.strip().isdigit()}
        if selected_ids:
            participants = [p for p in participants if p.get("user_id") in selected_ids]

    if not participants:
        raise HTTPException(status_code=404, detail="No participants found")

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=(105 * mm, 148 * mm))

    logo_left  = load_logo_with_transparency("assets/logo_left.png")
    logo_right = load_logo_with_transparency("assets/logo_right.png")

    for p in participants:
        _render_badge_page(c, p, logo_left, logo_right)

    c.save()
    buffer.seek(0)

    # Stamp when each of these badges was exported, so the "Choose Names"
    # picker can flag/skip people already exported today next time.
    # updated_at is explicitly self-assigned here to suppress its onupdate=
    # func.now() — otherwise SQLAlchemy's bulk update() re-evaluates onupdate
    # for every column *not* named in the values dict, silently bumping
    # updated_at for the whole batch (this previously stamped ~all 522
    # registrations to one identical timestamp and broke the admin
    # participant list's "most recently active" sort).
    reg_ids = [p["registration_id"] for p in participants]
    db.query(Registration).filter(Registration.id.in_(reg_ids)).update(
        {"badge_exported_at": datetime.utcnow(), "updated_at": Registration.updated_at},
        synchronize_session=False,
    )
    db.commit()

    safe_event_name = sanitize_filename(event.event)
    ascii_filename = f"{safe_event_name}_participant_badges.pdf"
    utf8_filename = urllib.parse.quote(ascii_filename)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={ascii_filename}; filename*=UTF-8''{utf8_filename}"
        },
    )


@router.get("/{event_id}/blank-role-badges/pdf")
async def download_blank_role_badges_pdf(
    event_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Generate blank A5 badges (no name, no QR) for support-staff roles that
    aren't individually registered — Ushers, Medical Team, Drivers, Media —
    with empty Name/Designation/Organization fields for on-site handwriting."""
    auth_dependency.secure_access("VIEW_EVENT", current_user["user_id"])

    event = db.query(Event).filter(Event.id == event_id, Event.deleted_at == None).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    roles = ["usher", "medical_staff", "driver", "media"]

    # A5 is the same ISO 216 aspect ratio as A6, so scaling up by the A5:A6
    # ratio reuses every hand-tuned A6 coordinate in _render_badge_page
    # as-is, instead of re-deriving the whole layout at a new size.
    A5_W, A5_H = 148 * mm, 210 * mm
    scale_x = A5_W / (105 * mm)
    scale_y = A5_H / (148 * mm)

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=(A5_W, A5_H))

    logo_left  = load_logo_with_transparency("assets/logo_left.png")
    logo_right = load_logo_with_transparency("assets/logo_right.png")

    for role_key in roles:
        p = {
            "registration_id": 0, "user_id": 0, "event_id": event_id,
            "title": "", "firstname": "", "middle_name": "", "lastname": "",
            "position": "", "organisation": "", "country": None,
            "event_name": event.event, "location": event.location or "",
            "event_theme": event.theme or "", "event_dates": _fmt_event_dates(event),
            "participation_role_raw": role_key, "paid": True, "has_proof": False,
        }
        # Each new page starts at an identity transform (showPage resets it),
        # so this scale call doesn't need a matching restoreState.
        c.scale(scale_x, scale_y)
        _render_badge_page(c, p, logo_left, logo_right, blank=True)

    c.save()
    buffer.seek(0)

    safe_event_name = sanitize_filename(event.event)
    filename = f"{safe_event_name}_blank_role_badges_A5.pdf"

    return StreamingResponse(
        buffer, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{event_id}/my-badge")
async def download_my_badge(
    event_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
):
    """Generate a badge PDF for the currently authenticated paid user."""
    event = get_object(event_id, db, Event)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    reg = db.query(Registration).filter(
        Registration.user_id == current_user["user_id"],
        Registration.event_id == event_id,
        Registration.deleted_at == None,
    ).first()

    if not reg:
        raise HTTPException(status_code=404, detail="You are not registered for this event")
    if not reg.paid:
        raise HTTPException(status_code=403, detail="Badge is only available after payment is confirmed")

    user = reg.user
    profile = user.user_profile[0] if user.user_profile else None
    country = reg.badge_country or (profile.country.country if profile and profile.country else None)
    organisation = reg.badge_organisation or (profile.organisation if profile else None)
    role_key = (
        reg.participation_role.name
        if hasattr(reg.participation_role, "name")
        else str(reg.participation_role).lower()
    )

    primary_color = (event.org_unit.primary_color or "#0095B6") if event.org_unit else "#0095B6"
    secondary_color = (event.org_unit.secondary_color or "#F7941D") if event.org_unit else "#F7941D"
    primary_rgb = hex_to_rgb(primary_color)
    secondary_rgb = hex_to_rgb(secondary_color)

    p = {
        "registration_id": reg.id,
        "event_id": event_id,
        "title": (reg.badge_prefix or (profile.title if profile else "")),
        "firstname": user.firstname,
        "middle_name": profile.middle_name if profile else "",
        "lastname": user.lastname,
        "position": (reg.badge_position or (profile.position if profile else "")),
        "organisation": organisation,
        "country": country,
        "participation_role": PARTICIPATION_ROLE_MAP.get(role_key, role_key),
        "event_name": event.event,
        "location": event.location or "",
        "event_theme": event.theme or "",
        "event_dates": _fmt_event_dates(event),
        "participation_role_raw": role_key,
        "paid": reg.paid,
    }

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=(105 * mm, 148 * mm))

    logo_left  = load_logo_with_transparency("assets/logo_left.png")
    logo_right = load_logo_with_transparency("assets/logo_right.png")

    _render_badge_page(c, p, logo_left, logo_right)

    c.save()
    buffer.seek(0)

    safe_name = sanitize_filename(f"{user.firstname}_{user.lastname}")
    ascii_filename = f"badge_{safe_name}.pdf"
    utf8_filename = urllib.parse.quote(ascii_filename)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={ascii_filename}; filename*=UTF-8''{utf8_filename}"
        },
    )


@router.get("/{event_id}/attendance")
async def get_event_attendance(
    event_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Admin: get all attendance records for an event."""
    auth_dependency.secure_access("VIEW_EVENT", current_user["user_id"])

    from models.models import EventAttendance
    records = (
        db.query(EventAttendance)
        .join(Registration, EventAttendance.registration_id == Registration.id)
        .filter(Registration.event_id == event_id)
        .order_by(EventAttendance.attendance_date.desc())
        .all()
    )

    result = []
    for a in records:
        reg = a.registration
        user = reg.user if reg else None
        profile = user.user_profile[0] if user and user.user_profile else None
        result.append({
            "id": a.id,
            "registration_id": a.registration_id,
            "attendance_date": a.attendance_date,
            "firstname": user.firstname if user else "",
            "lastname": user.lastname if user else "",
            "email": user.email if user else "",
            "organisation": profile.organisation if profile else "",
            "country": profile.country.country if profile and profile.country else "",
            "participation_role": reg.participation_role.name if reg else "",
            "paid": reg.paid if reg else False,
        })

    return {"total": len(result), "data": result}


@router.delete("/{event_id}/attendance")
async def reset_event_attendance(
    event_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Admin: delete all attendance records for an event."""
    auth_dependency.secure_access("VIEW_EVENT", current_user["user_id"])

    from models.models import EventAttendance
    records = (
        db.query(EventAttendance)
        .join(Registration, EventAttendance.registration_id == Registration.id)
        .filter(Registration.event_id == event_id)
        .all()
    )
    count = len(records)
    for r in records:
        db.delete(r)
    db.commit()
    return {"detail": f"Deleted {count} attendance record(s)"}


@router.get("/{event_id}/attendance/export")
async def export_event_attendance(
    event_id: int,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Admin: export attendance records as Excel."""
    auth_dependency.secure_access("VIEW_EVENT", current_user["user_id"])

    from models.models import EventAttendance
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    records = (
        db.query(EventAttendance)
        .join(Registration, EventAttendance.registration_id == Registration.id)
        .filter(Registration.event_id == event_id)
        .order_by(EventAttendance.attendance_date.asc())
        .all()
    )

    event = db.query(Event).filter(Event.id == event_id).first()
    event_name = event.event if event else f"Event {event_id}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ["#", "First Name", "Last Name", "Email", "Organisation", "Country", "Role", "Check-in Time", "Payment"]
    header_fill = PatternFill("solid", fgColor="1B3F6E")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, a in enumerate(records, 1):
        reg = a.registration
        user = reg.user if reg else None
        profile = user.user_profile[0] if user and user.user_profile else None
        checkin = a.attendance_date
        checkin_str = checkin.strftime("%Y-%m-%d %H:%M") if checkin else ""
        ws.append([
            i,
            user.firstname if user else "",
            user.lastname if user else "",
            user.email if user else "",
            profile.organisation if profile else "",
            profile.country.country if profile and profile.country else "",
            reg.participation_role.name if reg else "",
            checkin_str,
            "Paid" if (reg.paid if reg else False) else "Unpaid",
        ])

    col_widths = [4, 16, 16, 28, 26, 18, 16, 20, 10]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in event_name)[:40]
    filename = f"attendance_{safe_name}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


class PaymentReminderSchema(BaseModel):
    registration_ids: list[int] = []  # empty = send to ALL unpaid

@router.post("/{event_id}/send-payment-reminders")
async def send_payment_reminders(
    event_id: int,
    background_tasks: BackgroundTasks,
    current_user: user_dependency,
    body: PaymentReminderSchema = PaymentReminderSchema(),
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Admin: send payment reminder emails to selected (or all) unpaid registrants."""
    auth_dependency.secure_access("VIEW_EVENT", current_user["user_id"])

    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    q = (
        db.query(Registration)
        .join(User, Registration.user_id == User.id)
        .filter(
            Registration.event_id == event_id,
            Registration.paid == False,
            Registration.deleted_at == None,
            User.deleted_at == None,
        )
        .options(joinedload(Registration.user))
    )

    if body.registration_ids:
        q = q.filter(Registration.id.in_(body.registration_ids))

    targets = q.all()

    if not targets:
        return {"sent": 0, "message": "No matching unpaid registrations found."}

    import utils.mailer_util as mailer_util
    now = datetime.utcnow()
    for reg in targets:
        user = reg.user
        if user and user.email:
            mailer_util.payment_reminder_email(
                recipient_email=user.email,
                firstname=user.firstname or "Participant",
                event_name=event.event,
                payment_url="https://ecsahc.org/payment/",
                portal_url="https://events.ecsahc.org",
                background_tasks=background_tasks,
                db=db,
                sent_by_user_id=current_user["user_id"],
            )
            reg.reminder_sent_at = now

    db.commit()

    return {
        "sent": len(targets),
        "message": f"Payment reminder sent to {len(targets)} participant(s).",
    }


def _confirmed_participants(event_id: int, db: Session):
    """Paid OR proof-of-payment-uploaded registrants for an event, deduped by
    user — matches the "confirmed" definition already used for badges."""
    regs = (
        db.query(Registration)
        .join(User, Registration.user_id == User.id)
        .filter(
            Registration.event_id == event_id,
            Registration.deleted_at == None,
            User.deleted_at == None,
            or_(Registration.paid == True, Registration.payment_proof.isnot(None)),
        )
        .options(joinedload(Registration.user))
        .all()
    )
    by_email = {}
    for reg in regs:
        user = reg.user
        if not user or not user.email:
            continue
        key = user.email.lower()
        by_email.setdefault(key, {"user_id": user.id, "firstname": user.firstname or "Participant", "email": user.email})
    return by_email


class NotifyUpdateInfoSchema(BaseModel):
    deadline_label: str
    test_email: Optional[str] = None
    user_ids: Optional[list[int]] = None


@router.get("/{event_id}/update-info-notify-preview")
def update_info_notify_preview(
    event_id: int,
    current_user: user_dependency,
    deadline_label: str = Query(...),
    user_ids: Optional[str] = Query(None, description="Comma-separated user IDs to include"),
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Preview who would receive a 'please update your info' email (no emails sent)."""
    auth_dependency.secure_access("VIEW_EVENT", current_user["user_id"])
    from models.models import EmailLog
    import utils.mailer_util as _mailer

    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    by_email = _confirmed_participants(event_id, db)

    notified_emails = {
        row.recipient_email.lower()
        for row in db.query(EmailLog.recipient_email).filter(
            EmailLog.email_type == f"update_info_{event_id}", EmailLog.status == "sent",
        ).all()
    }

    to_send = [v for k, v in by_email.items() if k not in notified_emails]
    already_notified = [v for k, v in by_email.items() if k in notified_emails]

    # Filter by specific user IDs if provided
    if user_ids:
        selected_ids = {int(uid) for uid in user_ids.split(",") if uid.strip().isdigit()}
        if selected_ids:
            to_send = [v for v in to_send if v.get("user_id") in selected_ids]

    sample_name = to_send[0]["firstname"] if to_send else "Participant"
    email_preview_html = _mailer.templates.get_template("update_info_reminder_template.html").render(
        subject=f"Action Required: Please Verify Your Registration Details – {event.event}",
        firstname=sample_name,
        event_name=event.event,
        deadline_label=deadline_label,
        portal_url=_mailer.APP_BASE_URL,
        year=_mailer.YEAR,
    )

    return {
        "event_name": event.event,
        "to_send": to_send,
        "already_notified": already_notified,
        "total_recipients": len(by_email),
        "email_preview_html": email_preview_html,
    }


@router.post("/{event_id}/notify-update-info")
def notify_update_info(
    event_id: int,
    schema: NotifyUpdateInfoSchema,
    background_tasks: BackgroundTasks,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Email paid/confirmed participants asking them to log in and verify
    their profile details before the given deadline."""
    auth_dependency.secure_access("VIEW_EVENT", current_user["user_id"])
    from models.models import EmailLog
    import utils.mailer_util as _mailer

    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    email_type = f"update_info_{event_id}"
    by_email = _confirmed_participants(event_id, db)
    notified_emails = {
        row.recipient_email.lower()
        for row in db.query(EmailLog.recipient_email).filter(
            EmailLog.email_type == email_type, EmailLog.status == "sent",
        ).all()
    }
    jobs = [v for k, v in by_email.items() if k not in notified_emails]

    # Filter by specific user IDs if provided
    if schema.user_ids:
        selected_ids = set(schema.user_ids)
        jobs = [v for v in jobs if v.get("user_id") in selected_ids]

    if schema.test_email:
        jobs = jobs[:1] or [{"firstname": "Test", "email": schema.test_email}]
    elif not jobs:
        return {"sent": 0, "message": "No unnotified confirmed participants found for this event."}

    subject = f"Action Required: Please Verify Your Registration Details – {event.event}"

    messages = []
    for j in jobs:
        recipient = schema.test_email or j["email"]
        body = _mailer.templates.get_template("update_info_reminder_template.html").render(
            subject=subject,
            firstname=j["firstname"],
            event_name=event.event,
            deadline_label=schema.deadline_label,
            portal_url=_mailer.APP_BASE_URL,
            year=_mailer.YEAR,
        )
        messages.append({
            "recipient_email": recipient,
            "subject": subject,
            "body": body,
            "email_type": email_type if not schema.test_email else "update_info_test",
            "sent_by_user_id": current_user["user_id"],
        })

    background_tasks.add_task(_mailer.send_bulk_emails, messages, db)

    return {
        "sent": len(messages),
        "message": (
            f"Notification queued for {len(messages)} participant(s)."
            if not schema.test_email else "Test email queued."
        ),
    }


@router.get("/{event_id}/profile-changes")
def get_profile_changes(
    event_id: int,
    current_user: user_dependency,
    since: Optional[str] = Query(None, description="ISO datetime to filter changes after"),
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Return profile field changes made by participants of this event."""
    auth_dependency.secure_access("VIEW_EVENT", current_user["user_id"])
    from models.models import ActivityLog, User

    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    # Get user IDs registered for this event
    registered_user_ids = [
        reg.user_id for reg in event.registrations
        if reg.user_id and not reg.deleted_at
    ]
    if not registered_user_ids:
        return []

    # Query PROFILE_CHANGED logs for these users
    from sqlalchemy import and_
    query_filters = [
        ActivityLog.user_id.in_(registered_user_ids),
        ActivityLog.action == "PROFILE_CHANGED",
        ActivityLog.deleted_at == None,
    ]
    if since:
        try:
            since_dt = datetime.fromisoformat(since.replace("Z", "+00:00"))
            query_filters.append(ActivityLog.created_at >= since_dt)
        except ValueError:
            pass

    logs = (
        db.query(ActivityLog)
        .filter(and_(*query_filters))
        .order_by(ActivityLog.created_at.desc())
        .all()
    )

    # Build user name lookup
    user_ids = {log.user_id for log in logs}
    users = db.query(User).filter(User.id.in_(user_ids)).all()
    user_names = {u.id: f"{u.firstname or ''} {u.lastname or ''}".strip() or u.email for u in users}

    return [
        {
            "id": log.id,
            "user_id": log.user_id,
            "user_name": user_names.get(log.user_id, "Unknown"),
            "field": log.additional_data.get("label", log.additional_data.get("field", "")),
            "old_value": log.additional_data.get("old_value", ""),
            "new_value": log.additional_data.get("new_value", ""),
            "changed_at": log.created_at.isoformat() if log.created_at else None,
        }
        for log in logs
    ]


@router.post("/{event_id}/send-pending-reminder/{registration_id}")
async def send_single_pending_reminder(
    event_id: int,
    registration_id: int,
    background_tasks: BackgroundTasks,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Admin: send a payment reminder to a single pending registration."""
    auth_dependency.secure_access("ADMIN_DASHBOARD", current_user["user_id"])

    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    registration = db.query(Registration).filter(
        Registration.id == registration_id,
        Registration.event_id == event_id,
        Registration.deleted_at == None,
    ).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    import utils.mailer_util as mailer_util
    payment_url = f"https://events.ecsahc.org/payment/{event_id}/{registration_id}"
    if registration.user and registration.user.email:
        mailer_util.payment_reminder_email(
            recipient_email=registration.user.email,
            firstname=registration.user.firstname or "Participant",
            event_name=event.event,
            payment_url=payment_url,
            portal_url="https://events.ecsahc.org",
            background_tasks=background_tasks,
            db=db,
            sent_by_user_id=current_user["user_id"],
        )

    return {"message": "Reminder sent successfully."}


@router.post("/{event_id}/send-pending-bulk-reminders")
async def send_pending_bulk_reminders(
    event_id: int,
    background_tasks: BackgroundTasks,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Admin: send payment reminders to all pending (no proof uploaded) registrations."""
    auth_dependency.secure_access("ADMIN_DASHBOARD", current_user["user_id"])

    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    all_targets = (
        db.query(Registration)
        .filter(
            Registration.event_id == event_id,
            Registration.paid == False,
            Registration.payment_proof == None,
            Registration.deleted_at == None,
        )
        .all()
    )

    # Exclude accepted abstract authors — they get their own reminder from the abstracts page
    from sqlalchemy import text as _sql_text
    _author_rows = db.execute(_sql_text(
        "SELECT DISTINCT LOWER(aa.email) as email FROM abstract a "
        "JOIN abstract_author aa ON aa.abstract_id = a.id "
        "WHERE a.event_id = :eid AND a.status = 'accepted' AND aa.email IS NOT NULL"
    ), {"eid": event_id}).fetchall()
    excluded_emails = {row.email for row in _author_rows}

    targets = [
        t for t in all_targets
        if t.user and t.user.email and t.user.email.lower() not in excluded_emails
    ]

    if not targets:
        return {"sent": 0, "message": "No pending registrations found (abstract authors are excluded)."}

    import utils.mailer_util as mailer_util
    from starlette.templating import Jinja2Templates as _Jinja2
    _templates = _Jinja2(directory="templates")

    messages = []
    for reg in targets:
        if reg.user and reg.user.email:
            payment_url = f"https://events.ecsahc.org/payment/{event_id}/{reg.id}"
            subject = f"Action Required: Complete Your Payment – {event.event}"
            body = _templates.get_template("payment_reminder_template.html").render(
                subject=subject,
                firstname=reg.user.firstname or "Participant",
                event_name=event.event,
                payment_url=payment_url,
                portal_url="https://events.ecsahc.org",
                year=__import__("datetime").datetime.now().year,
            )
            messages.append({
                "recipient_email": reg.user.email,
                "subject": subject,
                "body": body,
                "email_type": "payment_reminder",
                "sent_by_user_id": current_user["user_id"],
            })

    background_tasks.add_task(mailer_util.send_bulk_emails, messages, db)

    return {
        "sent": len(messages),
        "excluded_abstract_authors": len(excluded_emails),
        "message": f"Payment reminder queued for {len(messages)} pending registration(s). {len(excluded_emails)} abstract author(s) excluded.",
    }


# ── Admin: look up user by email ──────────────────────────────────────────────
@router.get("/{event_id}/lookup-user")
async def lookup_user_by_email(
    event_id: int,
    email: str = Query(...),
    current_user: user_dependency = None,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Look up whether a user with the given email already exists."""
    auth_dependency.secure_access("ADMIN_DASHBOARD", current_user["user_id"])
    user = db.query(User).filter(User.email == email, User.deleted_at == None).first()
    if not user:
        return {"exists": False}

    # Check if already registered for this event
    reg = db.query(Registration).filter(
        Registration.user_id == user.id,
        Registration.event_id == event_id,
        Registration.deleted_at == None,
    ).first()

    return {
        "exists": True,
        "user_id": user.id,
        "firstname": user.firstname,
        "lastname": user.lastname,
        "email": user.email,
        "already_registered": bool(reg),
        "current_role": reg.participation_role.name if reg else None,
    }


# ── Admin: add (or re-register) a participant to an event ────────────────────
class AdminAddParticipantSchema(BaseModel):
    email: str
    firstname: str = ""
    lastname: str = ""
    participation_role: str
    send_invitation: bool = True
    payment_url: str = "https://ecsahc.org/payment/"
    portal_url: str = "https://events.ecsahc.org"


@router.post("/{event_id}/admin-add-participant")
async def admin_add_participant(
    event_id: int,
    body: AdminAddParticipantSchema,
    background_tasks: BackgroundTasks,
    current_user: user_dependency,
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Admin: find-or-create a user, register them to the event, optionally send invitation."""
    auth_dependency.secure_access("ADMIN_DASHBOARD", current_user["user_id"])

    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    # Validate role
    try:
        role_enum = ParticipationRole[body.participation_role]
    except KeyError:
        raise HTTPException(status_code=400, detail=f"Invalid participation role: {body.participation_role}")

    # ── 1. Find or create the user ────────────────────────────────────────────
    is_new_user = False
    temp_password = None

    user = db.query(User).filter(User.email == body.email, User.deleted_at == None).first()

    if not user:
        # Create new user with a temp password
        if not body.firstname or not body.lastname:
            raise HTTPException(
                status_code=400,
                detail="First name and last name are required for new users."
            )
        from passlib.hash import bcrypt as bcrypt_hash
        temp_password = auth_dependency.generate_random_password()
        hashed = bcrypt_hash.hash(temp_password)
        user = User(
            firstname=body.firstname,
            lastname=body.lastname,
            email=body.email,
            phone=None,
            hashed_password=hashed,
            verified=1,
        )
        db.add(user)
        db.commit()
        db.refresh(user)
        is_new_user = True
    else:
        # Update name if the admin supplied different values
        changed = False
        if body.firstname and body.firstname != user.firstname:
            user.firstname = body.firstname
            changed = True
        if body.lastname and body.lastname != user.lastname:
            user.lastname = body.lastname
            changed = True
        if changed:
            db.commit()

    # ── 2. Register (or update role if already registered) ────────────────────
    # NO_PAYMENT_ROLES (module-level, near BADGE_ROLE_COLORS) marks roles exempt
    # from payment — mark as paid automatically
    auto_paid = body.participation_role in NO_PAYMENT_ROLES

    existing_reg = db.query(Registration).filter(
        Registration.user_id == user.id,
        Registration.event_id == event_id,
        Registration.deleted_at == None,
    ).first()

    if existing_reg:
        existing_reg.participation_role = role_enum
        if auto_paid:
            existing_reg.paid = True
        db.commit()
        db.refresh(existing_reg)
        reg_id = existing_reg.id
        action = "updated"
    else:
        new_reg = Registration(
            user_id=user.id,
            event_id=event_id,
            participation_role=role_enum,
            paid=auto_paid,
        )
        db.add(new_reg)
        db.commit()
        db.refresh(new_reg)
        reg_id = new_reg.id
        action = "registered"

    # ── 3. Send email ─────────────────────────────────────────────────────────
    import utils.mailer_util as mailer_util

    # Build friendly role label (shared dict, also used by bulk_import_participants)
    role_label = _ROLE_LABELS.get(body.participation_role, body.participation_role)

    # Format event dates
    event_dates = None
    if event.start_date and event.end_date:
        def _fmt(d):
            return d.strftime("%-d %B %Y") if hasattr(d, "strftime") else str(d)
        event_dates = f"{_fmt(event.start_date)} – {_fmt(event.end_date)}"

    email_sent = False

    if is_new_user:
        # Brand-new user created inline: send welcome + event invitation with credentials
        mailer_util.event_invitation_email(
            recipient_email=user.email,
            firstname=user.firstname,
            event_name=event.event,
            participation_role=role_label,
            event_location=event.location,
            event_dates=event_dates,
            is_new_user=True,
            password=temp_password,
            no_payment=auto_paid,
            portal_url=body.portal_url,
            payment_url=body.payment_url,
            background_tasks=background_tasks,
            db=db,
            sent_by_user_id=current_user["user_id"],
        )
        user.credentials_sent = True
        db.commit()
        email_sent = True

    elif body.send_invitation:
        if not user.credentials_sent:
            # Existing user who has never received their credentials yet
            # (created silently via User Management). Issue a fresh password
            # so we can include it in the invitation.
            from passlib.hash import bcrypt as bcrypt_hash
            fresh_password = auth_dependency.generate_random_password()
            user.hashed_password = bcrypt_hash.hash(fresh_password)
            user.credentials_sent = True
            db.commit()
            db.refresh(user)
            mailer_util.event_invitation_email(
                recipient_email=user.email,
                firstname=user.firstname,
                event_name=event.event,
                participation_role=role_label,
                event_location=event.location,
                event_dates=event_dates,
                is_new_user=True,
                password=fresh_password,
                no_payment=auto_paid,
                portal_url=body.portal_url,
                payment_url=body.payment_url,
                background_tasks=background_tasks,
                db=db,
                sent_by_user_id=current_user["user_id"],
            )
        else:
            # Existing user who already has their credentials: send event-only invitation
            mailer_util.event_invitation_email(
                recipient_email=user.email,
                firstname=user.firstname,
                event_name=event.event,
                participation_role=role_label,
                event_location=event.location,
                event_dates=event_dates,
                is_new_user=False,
                password=None,
                no_payment=auto_paid,
                portal_url=body.portal_url,
                payment_url=body.payment_url,
                background_tasks=background_tasks,
                db=db,
                sent_by_user_id=current_user["user_id"],
            )
        email_sent = True

    return {
        "user_id": user.id,
        "registration_id": reg_id,
        "firstname": user.firstname,
        "lastname": user.lastname,
        "email": user.email,
        "is_new_user": is_new_user,
        "action": action,
        "email_sent": email_sent,
        "message": (
            f"{'New user created and registered' if is_new_user else f'Existing user {action}'}"
            f" to {event.event}."
            f"{' Invitation email sent.' if email_sent else ''}"
        ),
    }


# ── Admin: bulk-import participants from a spreadsheet ───────────────────────
_ROLE_LABELS = {
    "secretariat": "ECSA-HC Secretariat",
    "djcc": "DJCC Member",
    "moh": "Country Delegate (Ministry of Health)",
    "member_state": "Participant from ECSA Member States",
    "other_africa": "Participant from other African countries",
    "world": "International Participant",
    "student": "Student",
    "exhibitor": "Sponsor / Exhibitor",
    "participant": "Participant",
    "delegate": "Delegate",
    "presenter": "Presenter",
    "speaker": "Speaker",
    "sponsor": "Sponsor",
    "moderator": "Moderator",
    "local_secretariat": "Local Secretariat",
    "usher": "Usher",
    "driver": "Driver",
    "medical_staff": "Medical Staff",
}
_HONORIFICS = {"dr", "mr", "mrs", "ms", "prof", "rev", "eng", "hon"}
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _parse_full_name(raw_name: str):
    """Split a free-text name into (firstname, lastname, badge_prefix),
    stripping a leading honorific (Dr/Mr/Mrs/...) into badge_prefix. Returns
    (None, None, None) if nothing is left after removing the honorific."""
    parts = raw_name.split()
    honorific = None
    if parts and parts[0].strip(".").lower() in _HONORIFICS:
        honorific = parts[0].strip(".")
        parts = parts[1:]
    if not parts:
        return None, None, None
    firstname = parts[0]
    lastname = " ".join(parts[1:]) if len(parts) > 1 else parts[0]
    return firstname, lastname, honorific


def _parse_allowed_roles(participation_role: str):
    """participation_role is a comma-separated list of role keys (one entry
    when the admin picked a single role, several when the sheet mixes roles
    and a "Role" column will be matched per-row). Validates each key."""
    keys = [r.strip() for r in participation_role.split(",") if r.strip()]
    if not keys:
        raise HTTPException(status_code=400, detail="At least one participation role must be selected")
    for k in keys:
        if k not in ParticipationRole.__members__:
            raise HTTPException(status_code=400, detail=f"Invalid participation role: {k}")
    return keys


def _match_role_from_text(raw_text, allowed_role_keys):
    """Match a free-text Role/Category cell against the allowed role keys,
    checking both the enum key itself and its friendly label (case-
    insensitive, substring match either way). Returns the matched key, or
    None if nothing matches."""
    text = (raw_text or "").strip().lower()
    if not text:
        return None
    for key in allowed_role_keys:
        if text == key.lower():
            return key
        label = _ROLE_LABELS.get(key, "").lower()
        if label and (text == label or text in label or label in text):
            return key
    return None


@router.post("/{event_id}/bulk-import-participants")
async def bulk_import_participants(
    event_id: int,
    background_tasks: BackgroundTasks,
    current_user: user_dependency,
    file: UploadFile = File(...),
    participation_role: str = Form(...),
    send_invitation: bool = Form(True),
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Bulk-import participants from an uploaded spreadsheet (columns: Name,
    Title, Organization, Country, Email, payment_status — 'No' is ignored).
    Creates/registers each row like admin_add_participant, and returns a
    categorized report: imported, mismatches (imported but flagged), already
    registered, and rejected (with a reason).

    `participation_role` is a comma-separated list of role keys. With one
    role, it's applied to every row (the original behavior). With several
    (the sheet mixes multiple roles), each row's own "Role" column value is
    matched against the selected roles — rows with no match are rejected."""
    auth_dependency.secure_access("ADMIN_DASHBOARD", current_user["user_id"])
    import utils.mailer_util as mailer_util

    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    allowed_role_keys = _parse_allowed_roles(participation_role)

    from openpyxl import load_workbook
    contents = await file.read()
    try:
        wb = load_workbook(BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read file — please upload a valid .xlsx file")
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="File is empty")

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

    def find_col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    col_name = find_col("name")
    col_title = find_col("title")
    col_org = find_col("organization", "organisation")
    col_country = find_col("country")
    col_email = find_col("email")
    col_payment = find_col("payment_status", "payment status")
    col_role = find_col("role")

    if col_name is None or col_email is None:
        raise HTTPException(status_code=400, detail="File must have at least 'Name' and 'Email' columns")
    if len(allowed_role_keys) > 1 and col_role is None:
        raise HTTPException(
            status_code=400,
            detail="Multiple roles selected — the file needs a 'Role' column so each row can be matched to one of them",
        )

    def cell(row, idx):
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return str(v).strip() if v is not None else None

    # Some sheets carry a payment column with no header text at all — fall back
    # to the column immediately after Email if it holds recognizable payment
    # values (e.g. "PAID"), so an unlabeled column doesn't silently get ignored.
    if col_payment is None and col_email is not None:
        candidate = col_email + 1
        recognized = {"paid", "unpaid", "yes", "no", "true", "false", "1", "0", "y", "n"}
        if any(
            (v := cell(row, candidate)) is not None and v.strip().lower() in recognized
            for row in rows[1:]
        ):
            col_payment = candidate

    event_dates = None
    if event.start_date and event.end_date:
        def _fmt(d):
            return d.strftime("%-d %B %Y") if hasattr(d, "strftime") else str(d)
        event_dates = f"{_fmt(event.start_date)} – {_fmt(event.end_date)}"

    imported, mismatches, already_there, rejected = [], [], [], []

    for i, row in enumerate(rows[1:], start=2):
        raw_name = cell(row, col_name)
        raw_title = cell(row, col_title)
        raw_org = cell(row, col_org)
        raw_country = cell(row, col_country)
        raw_email = cell(row, col_email)
        raw_payment = cell(row, col_payment)
        raw_role = cell(row, col_role)

        if not raw_name and not raw_email:
            continue  # fully blank row

        row_ref = {"row": i, "name": raw_name, "email": raw_email}

        if not raw_name:
            rejected.append({**row_ref, "reason": "Missing name"})
            continue
        if not raw_email:
            rejected.append({**row_ref, "reason": "Missing email"})
            continue

        if len(allowed_role_keys) == 1:
            row_role_key = allowed_role_keys[0]
        else:
            row_role_key = _match_role_from_text(raw_role, allowed_role_keys)
            if row_role_key is None:
                rejected.append({
                    **row_ref,
                    "reason": f"Role '{raw_role or ''}' does not match any of the selected roles",
                })
                continue
        role_enum = ParticipationRole[row_role_key]
        role_label = _ROLE_LABELS.get(row_role_key, row_role_key)
        auto_paid_role = row_role_key in NO_PAYMENT_ROLES

        email_parts = [e.strip() for e in re.split(r"[;,]", raw_email) if e.strip()]
        if len(email_parts) > 1:
            rejected.append({
                **row_ref,
                "reason": f"Multiple emails in one row ({', '.join(email_parts)}) — split into separate rows and re-upload",
            })
            continue
        email = email_parts[0].lower()
        if not _EMAIL_RE.match(email):
            rejected.append({**row_ref, "reason": f"Invalid email format: {raw_email}"})
            continue

        firstname, lastname, honorific = _parse_full_name(raw_name)
        if firstname is None:
            rejected.append({**row_ref, "reason": "Name has no content after removing honorific"})
            continue

        payment_flag = (raw_payment or "").strip().lower()
        paid = payment_flag in {"paid", "yes", "true", "1", "y"}
        payment_note = None
        if raw_payment and not paid and payment_flag not in {"unpaid", "no", "false", "0", "n", ""}:
            payment_note = f"Unrecognized payment_status value '{raw_payment}' — treated as unpaid"

        # Dedicated badge_* columns (not `notes` — that's a free-text admin
        # scratchpad other features overwrite; badge rendering needs a stable source).
        badge_prefix = honorific or None
        badge_position = raw_title or None
        badge_organisation = raw_org or None
        badge_country = raw_country or None

        existing_user = db.query(User).filter(User.email == email, User.deleted_at == None).first()

        if existing_user:
            existing_reg = db.query(Registration).filter(
                Registration.user_id == existing_user.id,
                Registration.event_id == event_id,
                Registration.deleted_at == None,
            ).first()
            if existing_reg:
                already_there.append({
                    **row_ref, "email": email,
                    "existing_name": f"{existing_user.firstname} {existing_user.lastname}",
                    "reason": "Already registered for this event",
                })
                continue

            name_mismatch = (
                existing_user.firstname.strip().lower() != firstname.strip().lower()
                or existing_user.lastname.strip().lower() != lastname.strip().lower()
            )
            new_reg = Registration(
                user_id=existing_user.id, event_id=event_id,
                participation_role=role_enum, paid=paid or auto_paid_role,
                badge_prefix=badge_prefix, badge_position=badge_position, badge_organisation=badge_organisation,
                badge_country=badge_country,
            )
            db.add(new_reg)
            db.commit()
            db.refresh(new_reg)

            row_result = {**row_ref, "email": email, "registration_id": new_reg.id, "is_new_user": False}

            if send_invitation:
                if not existing_user.credentials_sent:
                    from passlib.hash import bcrypt as bcrypt_hash
                    fresh_password = auth_dependency.generate_random_password()
                    existing_user.hashed_password = bcrypt_hash.hash(fresh_password)
                    existing_user.credentials_sent = True
                    db.commit()
                    mailer_util.event_invitation_email(
                        recipient_email=existing_user.email, firstname=existing_user.firstname,
                        event_name=event.event, participation_role=role_label,
                        event_location=event.location, event_dates=event_dates,
                        is_new_user=True, password=fresh_password, no_payment=(paid or auto_paid_role),
                        portal_url="https://events.ecsahc.org", payment_url="https://ecsahc.org/payment/",
                        background_tasks=background_tasks, db=db, sent_by_user_id=current_user["user_id"],
                    )
                else:
                    mailer_util.event_invitation_email(
                        recipient_email=existing_user.email, firstname=existing_user.firstname,
                        event_name=event.event, participation_role=role_label,
                        event_location=event.location, event_dates=event_dates,
                        is_new_user=False, password=None, no_payment=(paid or auto_paid_role),
                        portal_url="https://events.ecsahc.org", payment_url="https://ecsahc.org/payment/",
                        background_tasks=background_tasks, db=db, sent_by_user_id=current_user["user_id"],
                    )

            reasons = []
            if name_mismatch:
                reasons.append(f"Sheet name '{raw_name}' differs from existing account name '{existing_user.firstname} {existing_user.lastname}'")
            if payment_note:
                reasons.append(payment_note)
            (mismatches if reasons else imported).append(
                {**row_result, "reason": "; ".join(reasons)} if reasons else row_result
            )
            continue

        # Brand-new user
        from passlib.hash import bcrypt as bcrypt_hash
        temp_password = auth_dependency.generate_random_password()
        hashed = bcrypt_hash.hash(temp_password)
        user = User(
            firstname=firstname, lastname=lastname, email=email,
            phone=None, hashed_password=hashed, verified=1,
        )
        db.add(user)
        db.commit()
        db.refresh(user)

        new_reg = Registration(
            user_id=user.id, event_id=event_id,
            participation_role=role_enum, paid=paid or auto_paid_role,
            badge_prefix=badge_prefix, badge_position=badge_position, badge_organisation=badge_organisation,
            badge_country=badge_country,
        )
        db.add(new_reg)
        db.commit()
        db.refresh(new_reg)

        row_result = {**row_ref, "email": email, "registration_id": new_reg.id, "is_new_user": True}

        if send_invitation:
            user.credentials_sent = True
            db.commit()
            mailer_util.event_invitation_email(
                recipient_email=user.email, firstname=user.firstname, event_name=event.event,
                participation_role=role_label, event_location=event.location, event_dates=event_dates,
                is_new_user=True, password=temp_password, no_payment=(paid or auto_paid_role),
                portal_url="https://events.ecsahc.org", payment_url="https://ecsahc.org/payment/",
                background_tasks=background_tasks, db=db, sent_by_user_id=current_user["user_id"],
            )

        if payment_note:
            mismatches.append({**row_result, "reason": payment_note})
        else:
            imported.append(row_result)

    return {
        "total_rows": len(rows) - 1,
        "imported": imported,
        "mismatches": mismatches,
        "already_there": already_there,
        "rejected": rejected,
        "summary": {
            "imported": len(imported),
            "mismatches": len(mismatches),
            "already_there": len(already_there),
            "rejected": len(rejected),
        },
    }


@router.post("/{event_id}/bulk-import-names-only")
async def bulk_import_names_only(
    event_id: int,
    current_user: user_dependency,
    file: UploadFile = File(...),
    participation_role: str = Form(...),
    db: Session = Depends(get_db),
    auth_dependency: Auth = Depends(get_auth_dependency),
):
    """Bulk-import badge-only participants with no email on file (e.g. ushers,
    drivers, medical staff, local secretariat support teams — no email column
    at all in these lists). Columns: Name (required), Position/Title,
    Organization, Category (folded into badge_position alongside Position).

    Each row still gets a real `User` row under the hood, because
    `Registration.user_id` is NOT NULL and unique per event — but the email
    is a generated, non-functional placeholder (no invite is ever sent, no
    password is ever usable), so these accounts can never be logged into.
    This lets the normal badge/participant-list code work completely
    unchanged. Duplicate rows (matched by first+last name within this event,
    since there's no email to match on) are skipped and reported.

    `participation_role` is a comma-separated list of role keys — same
    multi-role/"Role" column matching as bulk_import_participants."""
    auth_dependency.secure_access("ADMIN_DASHBOARD", current_user["user_id"])

    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    allowed_role_keys = _parse_allowed_roles(participation_role)

    from openpyxl import load_workbook
    contents = await file.read()
    try:
        wb = load_workbook(BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read file — please upload a valid .xlsx file")
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="File is empty")

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

    def find_col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    col_name = find_col("name")
    col_title = find_col("title", "position")
    col_org = find_col("organization", "organisation")
    col_category = find_col("category")
    col_role = find_col("role")

    if col_name is None:
        raise HTTPException(status_code=400, detail="File must have at least a 'Name' column")
    if len(allowed_role_keys) > 1 and col_role is None:
        raise HTTPException(
            status_code=400,
            detail="Multiple roles selected — the file needs a 'Role' column so each row can be matched to one of them",
        )

    def cell(row, idx):
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return str(v).strip() if v is not None else None

    # No email to match on for badge-only entries, so dedup by name within
    # this event — covers both real-account and previously placeholder rows.
    existing_names = {
        (u.firstname.strip().lower(), u.lastname.strip().lower())
        for u in db.query(User)
            .join(Registration, Registration.user_id == User.id)
            .filter(Registration.event_id == event_id, Registration.deleted_at == None)
            .all()
    }

    imported, already_there, rejected = [], [], []

    for i, row in enumerate(rows[1:], start=2):
        raw_name = cell(row, col_name)
        raw_title = cell(row, col_title)
        raw_org = cell(row, col_org)
        raw_category = cell(row, col_category)
        raw_role = cell(row, col_role)

        if not raw_name:
            continue  # fully blank row

        row_ref = {"row": i, "name": raw_name}

        firstname, lastname, honorific = _parse_full_name(raw_name)
        if firstname is None:
            rejected.append({**row_ref, "reason": "Name has no content after removing honorific"})
            continue

        name_key = (firstname.strip().lower(), lastname.strip().lower())
        if name_key in existing_names:
            already_there.append({**row_ref, "reason": "Already registered for this event"})
            continue

        if len(allowed_role_keys) == 1:
            row_role_key = allowed_role_keys[0]
        else:
            row_role_key = _match_role_from_text(raw_role, allowed_role_keys)
            if row_role_key is None:
                rejected.append({
                    **row_ref,
                    "reason": f"Role '{raw_role or ''}' does not match any of the selected roles",
                })
                continue
        role_enum = ParticipationRole[row_role_key]
        auto_paid_role = row_role_key in NO_PAYMENT_ROLES

        badge_prefix = honorific or None
        badge_position = raw_title or None
        if raw_category:
            badge_position = f"{badge_position} - {raw_category}" if badge_position else raw_category
        badge_organisation = raw_org or None

        from passlib.hash import bcrypt as bcrypt_hash
        temp_password = auth_dependency.generate_random_password()
        hashed = bcrypt_hash.hash(temp_password)
        placeholder_email = f"badge.{uuid.uuid4().hex[:12]}@no-login.ecsahc.internal"
        user = User(
            firstname=firstname, lastname=lastname, email=placeholder_email,
            phone=None, hashed_password=hashed, verified=1,
            credentials_sent=True,  # placeholder email — never attempt to send credentials
        )
        db.add(user)
        db.commit()
        db.refresh(user)

        new_reg = Registration(
            user_id=user.id, event_id=event_id,
            participation_role=role_enum, paid=auto_paid_role,
            badge_prefix=badge_prefix, badge_position=badge_position, badge_organisation=badge_organisation,
        )
        db.add(new_reg)
        db.commit()
        db.refresh(new_reg)

        existing_names.add(name_key)
        imported.append({**row_ref, "registration_id": new_reg.id})

    return {
        "total_rows": len(rows) - 1,
        "imported": imported,
        "already_there": already_there,
        "rejected": rejected,
        "summary": {
            "imported": len(imported),
            "already_there": len(already_there),
            "rejected": len(rejected),
        },
    }


# ── On-site registration (paid directly) ──────────────────────────────────────

@router.post("/onsite-register/")
def onsite_register(
    event_id: int = Form(...),
    firstname: str = Form(...),
    lastname: str = Form(...),
    designation: str = Form(None),
    organisation: str = Form(None),
    email: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    background_tasks: BackgroundTasks = None,
):
    """Register an attendee on-site with minimal info. Marked as paid immediately."""
    event = db.query(Event).filter(Event.id == event_id, Event.deleted_at == None).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    # Find or create user
    user = None
    if email:
        user = db.query(User).filter(User.email.ilike(email)).first()

    # Tracks whether we need to email login credentials below: set whenever
    # we generate a brand-new plaintext password (new account, or an
    # existing account that never had credentials issued).
    fresh_password = None

    if not user:
        # Create a new user with a properly hashed password
        from dependencies.auth_dependency import Auth as AuthCls
        auth_dep = AuthCls(db)
        fresh_password = auth_dep.generate_random_password()
        hashed = auth_dep.hash_password(fresh_password)

        user = User(
            firstname=firstname,
            lastname=lastname,
            email=email or f"onsite_{secrets.token_hex(8)}@event.local",
            phone=None,
            hashed_password=hashed,
            verified=bool(email),
            must_change_password=True,
        )
        db.add(user)
        db.flush()

        # Assign the "User" role
        from models.models import Role, UserRole
        role = db.query(Role).filter(Role.role == "User").first()
        if role:
            db.add(UserRole(user_id=user.id, role_id=role.id))

        # No UserProfile here — this crashed before (UserProfile has no
        # "designation" column, it's "position"; also country_id/title/
        # middle_name/gender are all NOT NULL and this minimal on-site form
        # never collects them). The Registration below already stores
        # designation/organisation via badge_position/badge_organisation,
        # which badge rendering and the admin participant list already read
        # as a fallback when there's no UserProfile — no data is lost by
        # skipping profile creation for a walk-in registrant.

        # Create account verification record
        from models.models import AccountVerification
        db.add(AccountVerification(
            user_id=user.id,
            verification_token=str(uuid.uuid4()),
            expires_at=datetime.utcnow() + timedelta(hours=1),
        ))
    elif email and not user.credentials_sent:
        # Existing account (e.g. created silently elsewhere) that has never
        # actually been sent login credentials — issue a fresh password now
        # so this on-site registration is the moment they get portal access.
        from dependencies.auth_dependency import Auth as AuthCls
        auth_dep = AuthCls(db)
        fresh_password = auth_dep.generate_random_password()
        user.hashed_password = auth_dep.hash_password(fresh_password)

    # Check for existing registration
    existing = db.query(Registration).filter(
        Registration.user_id == user.id,
        Registration.event_id == event_id,
        Registration.deleted_at == None,
    ).first()

    if existing:
        # Update to paid
        existing.paid = True
        existing.participation_role = ParticipationRole.participant
        if designation:
            existing.badge_position = designation
        if organisation:
            existing.badge_organisation = organisation
        db.commit()
        result = {"message": "Registration updated and marked as paid", "registration_id": existing.id, "user_id": user.id}
        is_new_registration = False
    else:
        # Create new registration (paid directly)
        registration = Registration(
            user_id=user.id,
            event_id=event_id,
            participation_role=ParticipationRole.participant,
            paid=True,
            badge_position=designation or "",
            badge_organisation=organisation or "",
        )
        db.add(registration)
        db.commit()
        db.refresh(registration)
        result = {
            "message": "Registered and marked as paid",
            "registration_id": registration.id,
            "user_id": user.id,
        }
        is_new_registration = True

    # Email login credentials whenever a real email was given — either a
    # brand-new account, or an existing one that never had credentials sent.
    # Repeat registrants who already have credentials just get a short
    # "you're registered" notice for this specific event.
    if email:
        import utils.mailer_util as mailer_util

        def _fmt(d):
            return d.strftime("%-d %B %Y") if hasattr(d, "strftime") else str(d)
        event_dates = f"{_fmt(event.start_date)} – {_fmt(event.end_date)}" if event.start_date and event.end_date else None
        role_label = _ROLE_LABELS.get("participant", "Participant")

        if fresh_password:
            user.credentials_sent = True
            db.commit()
            mailer_util.event_invitation_email(
                recipient_email=user.email, firstname=user.firstname, event_name=event.event,
                participation_role=role_label, event_location=event.location, event_dates=event_dates,
                is_new_user=True, password=fresh_password, no_payment=True,
                portal_url="https://events.ecsahc.org", payment_url="https://ecsahc.org/payment/",
                background_tasks=background_tasks, db=db,
            )
        elif is_new_registration:
            mailer_util.event_invitation_email(
                recipient_email=user.email, firstname=user.firstname, event_name=event.event,
                participation_role=role_label, event_location=event.location, event_dates=event_dates,
                is_new_user=False, password=None, no_payment=True,
                portal_url="https://events.ecsahc.org", payment_url="https://ecsahc.org/payment/",
                background_tasks=background_tasks, db=db,
            )

    return result


# ── QR Code PDF for on-site registration ──────────────────────────────────────

@router.get("/{event_id}/onsite-qr-pdf")
def generate_onsite_qr_pdf(
    event_id: int,
    db: Session = Depends(get_db),
):
    """Generate a downloadable A4 PDF with a QR code linking to the on-site registration page."""
    event = db.query(Event).filter(Event.id == event_id, Event.deleted_at == None).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors

    # Build the registration URL
    client_origin = os.getenv("CLIENT_ORIGIN", "https://events.ecsahc.org")
    reg_url = f"{client_origin}/onsite-register?event_id={event_id}"

    # Generate QR code
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=10, border=2)
    qr.add_data(reg_url)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="#1B3F6E", back_color="white").convert("RGB")

    # Save QR to buffer
    qr_buf = io.BytesIO()
    qr_img.save(qr_buf, format="PNG")
    qr_buf.seek(0)
    qr_reader = ImageReader(qr_buf)

    # Load ECSA logo
    try:
        logo_right = load_logo_with_transparency("assets/logo_right.png")
    except Exception:
        logo_right = None

    # Build PDF
    buf = io.BytesIO()
    w, h = A4
    c = canvas.Canvas(buf, pagesize=A4)

    # Background
    c.setFillColor(colors.HexColor("#F8FAFC"))
    c.rect(0, 0, w, h, fill=1, stroke=0)

    # Top coloured bar
    c.setFillColor(colors.HexColor("#1B3F6E"))
    c.rect(0, h - 2.2 * cm, w, 2.2 * cm, fill=1, stroke=0)

    # ECSA logo on top bar
    if logo_right:
        try:
            logo_h = 1.4 * cm
            logo_w = logo_h * 2.7
            c.drawImage(logo_right, w - RM - logo_w - 0.3 * cm, h - 1.9 * cm, width=logo_w, height=logo_h,
                        preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

    # Title text
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(w / 2, h - 1.5 * cm, "ON-SITE REGISTRATION")

    # Event name (strip ordinal suffixes to avoid tofu boxes)
    c.setFillColor(colors.HexColor("#1B3F6E"))
    c.setFont("Helvetica-Bold", 14)
    event_name = normalize_event_name(event.event) if event.event else ""
    if len(event_name) > 60:
        event_name = event_name[:57] + "..."
    c.drawCentredString(w / 2, h - 3.8 * cm, event_name)

    # Date & location
    c.setFont("Helvetica", 11)
    c.setFillColor(colors.HexColor("#475569"))
    from datetime import datetime as dt
    date_str = f"{event.start_date.strftime('%d %B %Y')} – {event.end_date.strftime('%d %B %Y')}" if event.start_date and event.end_date else ""
    c.drawCentredString(w / 2, h - 4.8 * cm, f"{date_str}  •  {event.location or ''}")

    # QR code (centered)
    qr_size = 9 * cm
    qr_x = (w - qr_size) / 2
    qr_y = h / 2 - qr_size / 2 + 1 * cm
    c.drawImage(qr_reader, qr_x, qr_y, width=qr_size, height=qr_size)

    # Instruction text below QR
    c.setFillColor(colors.HexColor("#1B3F6E"))
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(w / 2, qr_y - 1.2 * cm, "Scan to Register")

    c.setFont("Helvetica", 10)
    c.setFillColor(colors.HexColor("#64748B"))
    c.drawCentredString(w / 2, qr_y - 2.2 * cm, "Scan this QR code with your phone camera to register on-site.")

    # Bottom bar
    c.setFillColor(colors.HexColor("#1B3F6E"))
    c.rect(0, 0, w, 1.2 * cm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica", 8)
    c.drawCentredString(w / 2, 0.4 * cm, "ECSA-HC • East, Central and Southern Africa Health Community")

    c.save()
    buf.seek(0)

    safe_e = unicodedata.normalize("NFKD", event.event).encode("ascii", "ignore").decode("ascii")
    safe_e = re.sub(r"[^\w\s-]", "", safe_e).strip().replace(" ", "_")[:30] or "event"
    filename = f"onsite_registration_qr_{safe_e}.pdf"

    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"})
